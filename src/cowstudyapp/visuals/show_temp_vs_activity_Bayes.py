import pandas as pd
import numpy as np
import statsmodels.api as sm
# import statsmodels.formula.api as smf
from pymer4.models import Lmer
# from statsmodels.genmod.bayes_mixed_glm import BinomialBayesMixedGLM
# import matplotlib.pyplot as plt
# import seaborn as sns
import os
# from scipy import stats
import ephem
# from datetime import date, datetime


# Import necessary libraries for parallelization
import concurrent.futures
import multiprocessing

from cowstudyapp.utils import from_posix_col
from cowstudyapp.config import ConfigManager


import sys

# Set higher memory limits
try:
    import resource
    # Increase stack size (in bytes)
    resource.setrlimit(resource.RLIMIT_STACK, (2**29, -1))
except ImportError:
    # resource module not available on Windows
    # Use Windows-specific approach if needed
    pass

# For Windows systems, try to increase memory allocation
if sys.platform == 'win32':
    try:
        import ctypes
        # Allow Python to use more memory (Windows specific)
        kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
        kernel32.SetProcessWorkingSetSize(-1, -1, -1)
        print("Memory limits increased for Windows")
    except Exception as e:
        print(f"Could not optimize Windows memory: {e}")

# Configure threading for numpy/scipy
try:
    import numpy as np
    np.seterr(all='ignore')  # Ignore numerical warnings
    os.environ["OMP_NUM_THREADS"] = "4"  # Control OpenMP threads
    os.environ["MKL_NUM_THREADS"] = "4"  # Control MKL threads
    print("NumPy thread settings optimized")
except Exception as e:
    print(f"Could not configure NumPy threading: {e}")


# Set up R for better performance
def optimize_r_performance():
    try:
        import rpy2.robjects as ro
        # Use multiple cores for BLAS operations
        ro.r('library(parallel)')
        cores = max(1, multiprocessing.cpu_count() - 1)
        ro.r(f'options(mc.cores = {cores})')
        
        # Increase memory limit if available
        ro.r('memory.limit(size=16000)')  # Set to 16GB if available
        
        # Set BLAS/LAPACK threading
        ro.r('Sys.setenv(OMP_NUM_THREADS = 4)')  # Adjust based on your system
        
        print("R performance optimizations applied")
    except Exception as e:
        print(f"Could not optimize R performance: {e}")

# Call this function at the beginning of your script
optimize_r_performance()

class GrazingVersusTemperatureBayes:
    def __init__(self, config: ConfigManager):
        self.config = config
        # Set up Norris, MT observer
        self.observer = ephem.Observer()
        self.observer.lat = '45.575'
        self.observer.lon = '-111.625'
        self.observer.elevation = 1715
        self.observer.pressure = 0
        self.observer.horizon = '-0:34'
        # self._model_counter = []  # Track model numbers
        self.buffer = 1.5


    def _add_cols(self, df):
        df = self._add_suntime_cols(df)
        df = self._add_temp_col(df)
        df = self._add_meta_info(df)
        return df

    def _add_meta_info(self,df):

        meta_info_df = pd.read_excel(self.config.io.cow_info_path)
        meta_info_df["age"] = 2022 - meta_info_df['year_b']
        meta_info_df.set_index('collar_id',inplace=True)
        meta_info_df['BW_preg'] = meta_info_df['BW_preg']*0.4535924 # lbs to kg
        age_dict = meta_info_df['age'].to_dict()
        weight_dict = meta_info_df['BW_preg'].to_dict()

        df['age'] = df["ID"].map(age_dict)
        df['weight'] = df["ID"].map(weight_dict)

        df = df[df['weight'] > 0]
        return df


    def _add_suntime_cols(self,df):
        """Calculate sunrise/sunset times for all rows"""

        print("Calculating sunrise and sunset times...")
        
        # Make sure we have dates
        if 'mt' not in df.columns:
            df['mt'] = from_posix_col(df['posix_time'], self.config.analysis.timezone)
        
        # Extract date component
        df['date'] = df['mt'].dt.date
        
        # Get unique dates
        unique_dates = df['date'].unique()
        print(f"Calculating sun times for {len(unique_dates)} unique dates...")
        
        # Calculate sunrise/sunset for each unique date
        sun_times = {}
        for date in unique_dates:
            # Calculate sunrise time
            self.observer.date = date.strftime('%Y/%m/%d 00:00:00')
            sunrise = pd.to_datetime(str(self.observer.next_rising(ephem.Sun())))
            sunrise = sunrise.tz_localize('UTC').tz_convert(self.config.analysis.timezone)
            
            # Calculate sunset time
            self.observer.date = date.strftime('%Y/%m/%d 12:00:00')
            sunset = pd.to_datetime(str(self.observer.next_setting(ephem.Sun())))
            sunset = sunset.tz_localize('UTC').tz_convert(self.config.analysis.timezone)
            

    
            sun_times[date] = {
                'sunrise': sunrise - pd.Timedelta(hours=self.buffer),
                'sunset': sunset + pd.Timedelta(hours=self.buffer)
            }
        
        # Create dataframes for merging
        sunrise_df = pd.DataFrame([
            {'date': date, 'sunrise': times['sunrise']} 
            for date, times in sun_times.items()
        ])
        
        sunset_df = pd.DataFrame([
            {'date': date, 'sunset': times['sunset']} 
            for date, times in sun_times.items()
        ])

        # Merge sunrise times
        df = pd.merge(df, sunrise_df, on='date', how='left')
        
        # Merge sunset times
        df = pd.merge(df, sunset_df, on='date', how='left')

        df['is_daylight'] = ((df['mt'] >= df['sunrise']) & (df['mt'] <= df['sunset']))

        return df


    def _add_temp_col(self, df):
        """Add temperature data by joining with temperature dataset"""
        if 'temp' not in df.columns:
            if 'temperature_gps' in df.columns:
                df.rename(columns={'temperature_gps': 'temp'}, inplace=True)
            else:
                # Read temperature data
                temp_df = pd.read_csv(self.config.analysis.target_dataset)
                
                # Ensure we have the right columns for joining
                required_cols = ['device_id', 'posix_time', 'temperature_gps']
                if not all(col in temp_df.columns for col in required_cols):
                    raise ValueError(f"Temperature dataset must contain columns: {required_cols}")
                
                # Perform inner join
                df = pd.merge(
                    df,
                    temp_df[['device_id', 'posix_time', 'temperature_gps']],
                    left_on=['ID', 'posix_time'],
                    right_on=['device_id', 'posix_time'],
                    how='inner'
                )
                
                # Rename the temp column
                df.rename(columns={'temperature_gps': 'temp'}, inplace=True)
                
                # Drop the extra device_id column from the join
                df.drop('device_id', axis=1, inplace=True)
        

        return df


    def analyze_temperature_behavior_for_period(self, df, period, results=None):
        """
        Analyze the relationship between temperature and behavior states
        for a specific period (day or night)
        
        Parameters:
        -----------
        df : pandas DataFrame
            Data frame with behavior states and temperature data
        period : str
            'day' or 'night'
        results : dict, optional
            Results dictionary to update
            
        Returns:
        --------
        dict
            Dictionary of model results for each state in the specified period
        """
        if results is None:
            results = {}
            
        print(f"\n{'-'*80}")
        print(f"ANALYZING TEMPERATURE EFFECTS ON BEHAVIOR DURING {period.upper()}")
        print(f"{'-'*80}")
        
        # Try to optimize R memory usage
        try:
            import rpy2.robjects as ro
            ro.r('gc()')  # Run garbage collection in R
            ro.r('options(warn=-1)')  # Suppress R warnings
        except Exception as e:
            print(f"Could not configure R: {e}")

        # Print data summary
        print(f"\nData Summary for {period}:")
        print(f"Temperature range: {df['temp'].min():.1f} to {df['temp'].max():.1f}°C")
        print(f"Age range: {df['age'].min()} to {df['age'].max()} years")
        print(f"Weight range: {df['weight'].min():.1f} to {df['weight'].max():.1f} kg")
        print(f"Number of cows: {df['ID'].nunique()}")
        print(f"Total observations: {len(df)}")
        
        # Store original means and standard deviations for back-transformation
        means = {
            'temp': df['temp'].mean(),
            'age': df['age'].mean(),
            'weight': df['weight'].mean()
        }
        
        stds = {
            'temp': df['temp'].std(),
            'age': df['age'].std(),
            'weight': df['weight'].std()
        }
        
        # Create clean dataframe with only necessary columns
        clean_df = pd.DataFrame()
        clean_df['ID'] = df['ID']
        
        # Define model variables - use a simpler model focused on main effects
        # and the most important interactions
        selected_factors = [
            'temp_z',
            'age_z',
            'weight_z'
        ]
        
        selected_interactions = [
            'temp_z:age_z',
            'temp_z:weight_z',
            'age_z:weight_z'
        ]
        
        # Add z-standardized variables
        for factor in ['temp', 'age', 'weight']:
            z_col = f'{factor}_z'
            if z_col in selected_factors:
                clean_df[z_col] = (df[factor] - means[factor]) / stds[factor]
        
        # Create indicator variables for each state
        for state in self.config.analysis.hmm.states:
            clean_df[f'is_{state.lower()}'] = (df['predicted_state'] == state).astype(int)
        
        # Define fixed effects formula
        fixed_effects = ' + '.join(selected_factors)
        if selected_interactions:
            fixed_effects += ' + ' + ' + '.join(selected_interactions)
        
        # Define random effects structure - use simpler structure for faster convergence
        random_effects = "(1|ID)"  # Random intercepts for cow ID
        # random_effects = "(1 + temp_z|ID)"  # Random slopes for temperature by cow ID
        
        # Create formula template
        formula_template = 'is_{} ~ {} + {}'
        
        # Store standardization info for back-transformation
        results['standardization_info'] = {
            'means': means,
            'stds': stds
        }
        
        # Store model specification
        results['model_specification'] = {
            'selected_factors': selected_factors,
            'selected_interactions': selected_interactions
        }
        
        # Store clean data
        results['data'] = clean_df.copy()
        
        # Analyze each behavior state
        state_results = {}
        for state in self.config.analysis.hmm.states:
            print(f"\n{'-'*40}")
            print(f"ANALYZING {state.upper()} STATE - {period.upper()}")
            print(f"{'-'*40}")
            
            # Create formula for this state
            formula = formula_template.format(state.lower(), fixed_effects, random_effects)
            print(f"Model formula: {formula}")
            
            try:
                # Run R garbage collection before fitting
                import gc
                gc.collect()
                
                try:
                    import rpy2.robjects as ro
                    ro.r('gc()')
                except:
                    pass
                    
                # Fit the model
                model = Lmer(formula, data=clean_df, family='binomial')
                fit_result = model.fit()
                
                # Print model results
                print(f"\nModel Results for {state} ({period}):")
                print(fit_result)
                
                # Store results
                state_results[state] = {
                    'formula': formula,
                    'model': model,
                    'fit_result': fit_result
                }
                
                # Print interpretable results
                print(f"\nInterpretable Results for {state} ({period}):")
                
                # Baseline probability
                baseline_prob = fit_result.loc['(Intercept)', 'Prob']
                print(f"• Baseline probability of {state} during {period}: {baseline_prob:.3f} ({baseline_prob*100:.1f}%)")
                
                # Scaling factors for interpretable units
                scale_factors = {
                    'temp': 5.0,  # per 5°C
                    'age': 1.0,   # per year
                    'weight': 30.0  # per 30kg
                }
                
                # Interpret each effect
                for idx, row in fit_result.iterrows():
                    if idx == '(Intercept)':
                        continue
                        
                    # Interpret based on parameter type
                    if idx == 'temp_z':
                        orig_var = 'temp'
                        # Convert from z-score to original scale
                        orig_coef = row['Estimate'] * (scale_factors[orig_var] / stds[orig_var])
                        prob_effect = row['Prob'] - 0.5
                        prob_dir = "increase" if prob_effect > 0 else "decrease"
                        print(f"• Temperature effect ({period}): 5°C increase causes {abs(prob_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'age_z':
                        orig_var = 'age'
                        orig_coef = row['Estimate'] * (scale_factors[orig_var] / stds[orig_var])
                        prob_effect = row['Prob'] - 0.5
                        prob_dir = "increase" if prob_effect > 0 else "decrease"
                        print(f"• Age effect ({period}): 1 year increase causes {abs(prob_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'weight_z':
                        orig_var = 'weight'
                        orig_coef = row['Estimate'] * (scale_factors[orig_var] / stds[orig_var])
                        prob_effect = row['Prob'] - 0.5
                        prob_dir = "increase" if prob_effect > 0 else "decrease"
                        print(f"• Weight effect ({period}): 30kg increase causes {abs(prob_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'temp_z:age_z':
                        prob_effect = row['Prob'] - 0.5
                        effect_dir = "stronger" if prob_effect > 0 else "weaker"
                        print(f"• Temperature effect is {effect_dir} with increasing age ({period}) {row['Sig']}")
                    
                    else:
                        # Generic handler for other effects
                        prob_effect = row['Prob'] - 0.5
                        effect_dir = "positive" if prob_effect > 0 else "negative"
                        print(f"• {idx}: {effect_dir} interaction effect of {abs(prob_effect*100):.1f}% {row['Sig']}")
                
                # Print model fit statistics
                if hasattr(model, 'AIC'):
                    print(f"\nModel Fit Statistics:")
                    print(f"  AIC: {model.AIC:.1f}")
                    print(f"  BIC: {model.BIC:.1f}")
                
            except Exception as e:
                print(f"Error fitting model for {state} during {period}: {str(e)}")
                import traceback
                traceback.print_exc()
                state_results[state] = {'error': str(e)}
                
            # Clean up memory
            gc.collect()
        
        # Store all results for this period
        return state_results

    def analyze_temperature_behavior_glmm(self, df, results={}):
        """
        Analyze the relationship between temperature and behavior states
        using GLMM with Bayesian inference
        
        Parameters:
        -----------
        df : pandas DataFrame
            Data frame with behavior states and temperature data
        results : dict
            Results dictionary to update
            
        Returns:
        --------
        dict
            Dictionary of model results for each state
        """
        print("\n" + "="*80)
        print(f"ANALYZING TEMPERATURE EFFECTS ON BEHAVIOR WITH DAY/NIGHT AS PREDICTOR")
        print("="*80)
        
        # Try to optimize R memory usage
        try:
            import rpy2.robjects as ro
            ro.r('gc()')  # Run garbage collection in R
            ro.r('options(warn=-1)')  # Suppress R warnings
        except Exception as e:
            print(f"Could not configure R: {e}")


        # Ensure we have the necessary columns
        required_cols = ['ID', 'predicted_state', 'temp', 'is_daylight', 'age', 'weight']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            print(f"Error: Missing required columns: {missing_cols}")
            return None
        
        # Print data summary
        print(f"\nData Summary:")
        print(f"Temperature range: {df['temp'].min():.1f} to {df['temp'].max():.1f}°C")
        print(f"Age range: {df['age'].min()} to {df['age'].max()} years")
        print(f"Weight range: {df['weight'].min():.1f} to {df['weight'].max():.1f} kg")
        print(f"Number of cows: {df['ID'].nunique()}")
        print(f"Total observations: {len(df)}")
        print(f"Day observations: {df['is_daylight'].sum()} ({df['is_daylight'].mean()*100:.1f}%)")
        print(f"Night observations: {(~df['is_daylight']).sum()} ({(~df['is_daylight']).mean()*100:.1f}%)")
        
        # Store original means and standard deviations for back-transformation
        means = {
            'temp': df['temp'].mean(),
            'age': df['age'].mean(),
            'weight': df['weight'].mean()
        }
        
        stds = {
            'temp': df['temp'].std(),
            'age': df['age'].std(),
            'weight': df['weight'].std()
        }
        
        selected_factors = [
            'temp_z',
            'age_z',
            'weight_z',
            'day'
        ]
        
        selected_interactions = [
            'temp_z:age_z',
            'temp_z:weight_z',
            'age_z:weight_z',
            'day:age_z',
            'day:weight_z',
            'day:temp_z'
        ]
        

        clean_df = pd.DataFrame()
        clean_df['ID'] = df['ID']
        clean_df['day'] = ~df['is_daylight'].astype(int)
        # clean_df = df[['ID', 'predicted_state', 'is_daylight']]
        
        # Add z-standardized variables based on selected factors
        if 'temp_z' in selected_factors:
            clean_df['temp_z'] = (df['temp'] - means['temp']) / stds['temp']
        
        if 'age_z' in selected_factors:
            clean_df['age_z'] = (df['age'] - means['age']) / stds['age']
        
        if 'weight_z' in selected_factors:
            clean_df['weight_z'] = (df['weight'] - means['weight']) / stds['weight']

        # Create indicator variables for each state
        for state in self.config.analysis.hmm.states:
            clean_df[f'is_{state.lower()}'] = (df['predicted_state'] == state).astype(int)

        # Delete original df to free memory
        del df
        import gc
        gc.collect()  # Run Python garbage collection
        

        # Define fixed effects formula
        fixed_effects = ' + '.join(selected_factors)
        if selected_interactions:
            fixed_effects += ' + ' + ' + '.join(selected_interactions)
        
        # Define random effects structure
        # random_effects = "(1|ID) + (1|day)"  # Random intercepts for cow ID
        # random_effects = "(1 + temp_z|ID)"
        random_effects = "(1|ID)"
        
        # Create complete formula template
        formula_template = 'is_{} ~ {} + {}'
        
        # Store the standardization info for back-transformation
        results['standardization_info'] = {
            'means': means,
            'stds': stds
        }
        
        # Store the selected factors for reference
        results['model_specification'] = {
            'selected_factors': selected_factors,
            'selected_interactions': selected_interactions
        }
        
        # Store a reference copy of the data (just the needed columns)
        results['data'] = clean_df.copy()
        
        # Analyze each behavior state
        state_results = {}
        for state in self.config.analysis.hmm.states:
            print(f"\n{'-'*40}")
            print(f"ANALYZING {state.upper()} STATE")
            print(f"{'-'*40}")
            
            # Complete formula for this state
            formula = formula_template.format(state.lower(), fixed_effects, random_effects)
            print(f"Model formula: {formula}")
            
            try:
                # Run R garbage collection before each model fit
                import rpy2.robjects as ro
                ro.r('gc()')
                # Fit the model
                model = Lmer(formula, data=clean_df, family='binomial')
                fit_result = model.fit()
                
                # Print the model results
                print("\nModel Results (z-standardized variables):")
                print(fit_result)
                
                # Store model results for this state
                state_results[state] = {
                    'formula': formula,
                    'model': model,
                    'fit_result': fit_result
                }
                
                # Print interpretable results (back-transformed)
                print("\nInterpretable Results (original scale):")
                
                # Intercept (baseline probability)
                intercept = fit_result.loc['(Intercept)', 'Estimate']
                baseline_prob = fit_result.loc['(Intercept)', 'Prob']
                print(f"  • Baseline probability of {state} at average predictors: {baseline_prob:.3f} ({baseline_prob*100:.1f}%)")
                
                # Day vs Night effect
                if 'day' in fit_result.index:
                    day_effect = fit_result.loc['day', 'Prob'] - 0.5
                    day_dir = "higher" if day_effect > 0 else "lower"
                    print(f"  • Day vs. Night: Probability is {abs(day_effect*100):.1f}% {day_dir} during day {fit_result.loc['day', 'Sig']}")

                # Scaling factors for interpretable units
                scale_factors = {
                    'temp': 5.0,  # per 5°C
                    'age': 1.0,          # per year
                    'weight': 30.0       # per 30kg
                }
                
                # Interpret each effect
                for idx, row in fit_result.iterrows():
                    if idx in ['(Intercept)', 'day']:
                        continue  # Already handled intercept and day effect
                        
                    # Get standardized coefficient and interpret
                    z_coef = row['Estimate']
                
                    # Format effect description based on parameter type
                    if idx == 'temp_z':
                        orig_var = 'temp'
                        orig_coef = z_coef * (scale_factors[orig_var] / stds[orig_var])
                        prob_effect = row['Prob'] - 0.5
                        prob_dir = "increase" if prob_effect > 0 else "decrease"
                        print(f"  • Temperature effect (day): 5°C increase causes {abs(prob_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'age_z':
                        orig_var = 'age'
                        orig_coef = z_coef * (scale_factors[orig_var] / stds[orig_var])
                        prob_effect = row['Prob'] - 0.5
                        prob_dir = "increase" if prob_effect > 0 else "decrease"
                        print(f"  • Age effect: 1 year increase causes {abs(prob_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'temp_z:day':
                        # This shows how temperature effect differs between day and night
                        prob_effect = row['Prob'] - 0.5
                        if prob_effect > 0:
                            print(f"  • Temperature effect is {abs(prob_effect*100):.1f}% stronger during day {row['Sig']}")
                        else:
                            print(f"  • Temperature effect is {abs(prob_effect*100):.1f}% weaker during day {row['Sig']}")
                    
                    elif idx == 'temp_z:age_z':
                        # Interaction between temperature and age
                        prob_effect = row['Prob'] - 0.5
                        effect_dir = "stronger" if prob_effect > 0 else "weaker"
                        print(f"  • Temperature effect is {effect_dir} with increasing age {row['Sig']}")
                    
                    else:
                        # Generic handler for other interactions
                        prob_effect = row['Prob'] - 0.5
                        effect_dir = "positive" if prob_effect > 0 else "negative"
                        print(f"  • {idx}: {effect_dir} interaction effect of {abs(prob_effect*100):.1f}% {row['Sig']}")
                
                # Print model fit statistics
                if hasattr(model, 'AIC'):
                    print(f"\nModel Fit Statistics:")
                    print(f"  AIC: {model.AIC:.1f}")
                    print(f"  BIC: {model.BIC:.1f}")
                    print(f"  Log-Likelihood: {model.logLike:.1f}")
                
            except Exception as e:
                print(f"Error fitting model for {state}: {str(e)}")
                import traceback
                traceback.print_exc()
                state_results[state] = {'error': str(e)}
        
            import gc
            gc.collect()

        # Store results
        results['combined'] = state_results
        
        return results

    # def _create_interpretable_table(self, results, format_type="unicode"):
    #     """
    #     Create an interpretable table focusing on probability changes
    #     while maintaining key statistical information
    #     """
    #     # Prepare the table data
    #     table_data = []
        
    #     # Define scaling information for explanatory text
    #     scaling_note = (
    #         "Note on interpretations:\n"
    #         "• Probability Change: The absolute change in the behavior's probability for the given change in predictor\n"
    #         "• Model coefficients are standardized for modeling but results are shown in interpretable units:\n"
    #         "  - Temperature effects are per 5°C change\n"
    #         "  - Age effects are per year change\n"
    #         "  - Weight effects are per 30kg change\n"
    #         "• Day/Night effects show how behaviors differ between day and night periods\n"
    #         "• Interactions with 'day' show how other effects (temperature, age) differ between day and night"
    #     )
        
    #     # Add header row with more intuitive column names
    #     header = ["State", "Parameter", 
    #             "Probability Change", 
    #             "Original Coefficient", "Odds Ratio", 
    #             "p-value", "Significance"]
    #     table_data.append(header)
        
    #     # Define parameter interpretations for better readability
    #     param_interpretations = {
    #         '(Intercept)': 'Baseline at average predictors (day)',
    #         'temp_z': 'Temperature effect (5°C increase)',
    #         'age_z': 'Age effect (1-year increase)',
    #         'weight_z': 'Weight effect (30kg increase)',
    #         'day': 'Day vs. Night effect',
    #         'temp_z:age_z': 'Temperature × Age interaction',
    #         'temp_z:weight_z': 'Temperature × Weight interaction',
    #         'age_z:weight_z': 'Age × Weight interaction',
    #         'day:temp_z': 'Temperature effect difference in day vs. night',
    #         'day:age_z': 'Age effect difference in day vs. night',
    #         'day:weight_z': 'Weight effect difference in day vs. night'
    #     }
        
    #     # Skip non-result keys
    #     skip_keys = ['data', 'standardization_info', 'model_specification']
        
    #     # Process each state from the combined results
    #     for state, state_results in results.get('combined', {}).items():
    #         if 'fit_result' not in state_results:
    #             continue
                
    #         fit_result = state_results['fit_result']
            
    #         # Add a header row for this state
    #         table_data.append([
    #             state.capitalize(),
    #             "=== Model Results ===",
    #             "",
    #             "",
    #             "",
    #             "",
    #             ""
    #         ])
            
    #         # Process each parameter
    #         for idx, row in fit_result.iterrows():
    #             # Get parameter name, using friendly names if available
    #             param_name = param_interpretations.get(idx, idx)
                
    #             # Get coefficient values
    #             estimate = row['Estimate']
    #             odds_ratio = row['OR']
                
    #             # Format the probability effect based on parameter type
    #             if idx == '(Intercept)':
    #                 # For intercept, show the baseline probability
    #                 baseline_prob = row['Prob']
    #                 prob_formatted = f"Baseline: {baseline_prob*100:.1f}%"
    #                 orig_coef = estimate  # No transformation for intercept
    #             elif idx == 'day':
    #                 # Day vs Night effect
    #                 prob_effect = row['Prob'] - 0.5
    #                 direction = "+" if prob_effect > 0 else "-"
    #                 prob_formatted = f"{direction}{abs(prob_effect*100):.1f}%"
    #                 orig_coef = estimate  # Day is binary, no transformation needed
    #             elif idx == 'day:temp_z':
    #                 # How temperature effect differs between day and night
    #                 prob_effect = row['Prob'] - 0.5
    #                 if prob_effect > 0:
    #                     prob_formatted = f"Stronger in day by {abs(prob_effect*100):.1f}%"
    #                 else:
    #                     prob_formatted = f"Weaker in day by {abs(prob_effect*100):.1f}%"
    #                 orig_coef = estimate  # Keep interaction coefficient as is
    #             elif idx == 'day:age_z':
    #                 # How age effect differs between day and night
    #                 prob_effect = row['Prob'] - 0.5
    #                 if prob_effect > 0:
    #                     prob_formatted = f"Stronger in day by {abs(prob_effect*100):.1f}%"
    #                 else:
    #                     prob_formatted = f"Weaker in day by {abs(prob_effect*100):.1f}%"
    #                 orig_coef = estimate
    #             elif idx == 'day:weight_z':
    #                 # How weight effect differs between day and night
    #                 prob_effect = row['Prob'] - 0.5
    #                 if prob_effect > 0:
    #                     prob_formatted = f"Stronger in day by {abs(prob_effect*100):.1f}%"
    #                 else:
    #                     prob_formatted = f"Weaker in day by {abs(prob_effect*100):.1f}%"
    #                 orig_coef = estimate
    #             elif idx == 'temp_z:age_z':
    #                 # Temperature and age interaction
    #                 prob_effect = row['Prob'] - 0.5
    #                 if prob_effect > 0:
    #                     prob_formatted = f"Temp. effect stronger with age (+{abs(prob_effect*100):.1f}%)"
    #                 else:
    #                     prob_formatted = f"Temp. effect weaker with age (-{abs(prob_effect*100):.1f}%)"
    #                 orig_coef = estimate
    #             elif idx in ['temp_z', 'age_z', 'weight_z']:
    #                 # Main effects
    #                 prob_effect = row['Prob'] - 0.5
    #                 direction = "+" if prob_effect > 0 else "-"
    #                 prob_formatted = f"{direction}{abs(prob_effect*100):.1f}%"
                    
    #                 # Back-transform to original scale
    #                 var_name = idx.replace('_z', '')
    #                 scale_factor = 5.0 if var_name == 'temp' else (1.0 if var_name == 'age' else 30.0)
    #                 std = results['standardization_info']['stds'][var_name]
    #                 orig_coef = estimate * (scale_factor / std)
    #             else:
    #                 # Other interactions
    #                 prob_effect = row['Prob'] - 0.5
    #                 direction = "+" if prob_effect > 0 else "-"
    #                 prob_formatted = f"{direction}{abs(prob_effect*100):.1f}%"
    #                 orig_coef = estimate  # Keep standardized
                
    #             # Format p-value and significance
    #             p_value = row['P-val']
                
    #             # Format p-value with scientific notation for very small values
    #             if p_value < 0.0001:
    #                 p_value_formatted = "<0.0001"
    #             else:
    #                 p_value_formatted = f"{p_value:.4f}"
                    
    #             significance = row['Sig']
                
    #             # Add the row to the table
    #             table_data.append([
    #                 "",  # State (only in header)
    #                 param_name,
    #                 prob_formatted,
    #                 f"{orig_coef:.4f}",
    #                 f"{odds_ratio:.4f}",
    #                 p_value_formatted,
    #                 significance
    #             ])
            
    #         # Add model fit statistics if available
    #         if hasattr(state_results.get('model', {}), 'AIC'):
    #             fit_row = [
    #                 "",
    #                 "Model Fit Statistics",
    #                 "",
    #                 f"AIC: {state_results['model'].AIC:.1f}",
    #                 f"BIC: {state_results['model'].BIC:.1f}",
    #                 "",
    #                 ""
    #             ]
    #             table_data.append(fit_row)
                
    #         # Add an empty row for spacing
    #         table_data.append([""] * len(header))
        
    #     # Format and return the table
    #     if format_type == "unicode":
    #         from tabulate import tabulate
    #         table_str = tabulate(table_data, headers="firstrow", tablefmt="fancy_grid")
    #         return table_str + "\n\n" + scaling_note
    #     else:
    #         # Create a DataFrame for Excel export
    #         df = pd.DataFrame(table_data[1:], columns=table_data[0])
    #         return df


    def _create_comparison_table(self, results, format_type="unicode"):
        """
        Create a table comparing day vs night behavior effects in the requested format
        
        Parameters:
        -----------
        results : dict
            Dictionary with day and night results
        format_type : str
            'unicode' or 'excel'
            
        Returns:
        --------
        str or DataFrame
            Formatted table
        """
        # Prepare table data
        table_data = []
        
        # Explanatory note
        scaling_note = (
            "Note on interpretations:\n"
            "• Effects are shown separately for day and night periods\n"
            "• Probability Change: The absolute change in behavior probability\n"
            "• Temperature effects are per 5°C change\n"
            "• Age effects are per year change\n"
            "• Weight effects are per 30kg change\n"
            "• Baseline shows probability at average temperature/age/weight"
        )
        
        # Create header
        header = ["", "Parameter", "Behavior State", 
                "Probability Change", "Original Coefficient", 
                "Odds Ratio", "p-value", "Significance"]
        table_data.append(header)
        
        # Parameter labels
        param_labels = {
            '(Intercept)': 'Baseline at average predictors',
            'temp_z': 'Temperature effect (5°C increase)',
            'age_z': 'Age effect (1-year increase)',
            'weight_z': 'Weight effect (30kg increase)',
            'temp_z:age_z': 'Temperature × Age interaction',
            'temp_z:weight_z': 'Temperature × Weight interaction',
            'age_z:weight_z': 'Age × Weight interaction'
        }
        
        # Process each time period (day, night)
        for period in ['day', 'night']:
            if period not in results:
                continue
                
            # Add period header
            table_data.append([
                period.capitalize(),
                "",
                "",
                "",
                "",
                "",
                "",
                ""
            ])
            
            # Process each parameter
            for param in param_labels:
                param_label = param_labels[param]
                
                # Add parameter row
                table_data.append([
                    "",
                    param_label,
                    "",
                    "",
                    "",
                    "",
                    "",
                    ""
                ])
                
                # Add state-specific rows for this parameter
                for state in self.config.analysis.hmm.states:
                    # Get results for this state if available
                    state_result = None
                    if period in results and state in results[period]:
                        state_results = results[period][state]
                        if 'fit_result' in state_results and param in state_results['fit_result'].index:
                            state_result = state_results['fit_result'].loc[param]
                    
                    if state_result is not None:
                        # Format values
                        if param == '(Intercept)':
                            # For baseline, show actual probability
                            prob = state_result['Prob']
                            prob_text = f"{prob*100:.1f}%"
                        else:
                            # For effects, show change in probability
                            prob_effect = state_result['Prob'] - 0.5
                            direction = "+" if prob_effect > 0 else ""
                            prob_text = f"{direction}{prob_effect*100:.1f}%"
                        
                        # Format coefficient based on parameter type
                        if param in ['temp_z', 'age_z', 'weight_z']:
                            # Back-transform standardized coefficients
                            var_name = param.replace('_z', '')
                            std = results.get('standardization_info', {}).get('stds', {}).get(var_name, 1.0)
                            
                            # Scale factors for interpretable units
                            scale_factor = 5.0 if var_name == 'temp' else (1.0 if var_name == 'age' else 30.0)
                            orig_coef = state_result['Estimate'] * (scale_factor / std)
                        else:
                            # Keep other coefficients as is
                            orig_coef = state_result['Estimate']
                        
                        # Format odds ratio
                        odds_ratio = state_result['OR']
                        
                        # Format p-value with appropriate precision
                        p_val = state_result['P-val']
                        if p_val < 0.0001:
                            p_val_text = "<0.0001"
                        else:
                            p_val_text = f"{p_val:.4f}"
                        
                        # Get significance indicator
                        sig = state_result['Sig']
                        
                        # Add data row
                        table_data.append([
                            "",
                            "",
                            state.capitalize(),
                            prob_text,
                            f"{orig_coef:.4f}",
                            f"{odds_ratio:.4f}",
                            p_val_text,
                            sig
                        ])
                    else:
                        # Add empty row when no results available
                        table_data.append([
                            "",
                            "",
                            state.capitalize(),
                            "N/A",
                            "N/A",
                            "N/A",
                            "N/A",
                            "N/A"
                        ])
                
                # Add model fit statistics for the last parameter
                if param == list(param_labels.keys())[-1]:
                    table_data.append([
                        "",
                        "Model Fit Statistics",
                        "",
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                    
                    for state in self.config.analysis.hmm.states:
                        aic = "N/A"
                        bic = "N/A"
                        
                        if period in results and state in results[period]:
                            model = results[period][state].get('model')
                            if model is not None and hasattr(model, 'AIC'):
                                aic = f"AIC: {model.AIC:.1f}"
                                bic = f"BIC: {model.BIC:.1f}"
                        
                        table_data.append([
                            "",
                            "",
                            state.capitalize(),
                            "",
                            aic,
                            bic,
                            "",
                            ""
                        ])
        
        # Format and return table
        if format_type == "unicode":
            from tabulate import tabulate
            table_str = tabulate(table_data, headers="firstrow", tablefmt="fancy_grid")
            return table_str + "\n\n" + scaling_note
        else:
            # Create DataFrame for Excel
            df = pd.DataFrame(table_data[1:], columns=table_data[0])
            return df


    def _export_table_to_excel(self, results, filename=None):
        """
        Export the interpretable table to Excel
        
        Parameters:
        -----------
        results : dict
            Dictionary of model results
        filename : str, optional
            Path to save Excel file, defaults to temp_behavior_table.xlsx in visuals folder
        """
        # Get the table as a DataFrame
        df = self._create_comparison_table(results, format_type="excel")

        # Set default filename if not provided
        if filename is None:
            filename = os.path.join(self.config.visuals.visuals_root_path, 'temp_behavior_table.xlsx')
        
        # Create a new Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Effects"
        
        # Write the DataFrame to Excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        period_font = Font(bold=True, color="FFFFFF")
        period_fill = PatternFill(start_color="8CB4E2", end_color="8CB4E2", fill_type="solid")
        param_font = Font(bold=True)
        param_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border
        
        # Apply formatting to all rows
        for row_idx, row in enumerate(ws.rows, 1):
            if row_idx == 1:  # Skip header row
                continue
                
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
            
            first_cell_value = row[0].value
            
            # Format period headers (Day, Night)
            if first_cell_value in ['Day', 'Night']:
                for cell in row:
                    cell.font = period_font
                    cell.fill = period_fill
                    cell.alignment = centered_alignment
            
            # Format parameter rows
            elif row[1].value in ['Baseline at average predictors', 
                                'Temperature effect (5°C increase)',
                                'Age effect (1-year increase)',
                                'Weight effect (30kg increase)',
                                'Temperature × Age interaction',
                                'Temperature × Weight interaction',
                                'Age × Weight interaction',
                                'Model Fit Statistics']:
                for cell in row:
                    cell.font = param_font
                    cell.fill = param_fill
            
            # Apply alternating colors to state rows
            elif row[2].value in ['Grazing', 'Resting', 'Traveling']:
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)
        print(f"Exported table to {filename}")




    def _create_comparison_table_compact(self, results, format_type="unicode"):
        """
        Create a more compact table comparing day vs night behavior effects
        
        Parameters:
        -----------
        results : dict
            Dictionary with day and night results
        format_type : str
            'unicode' or 'excel'
            
        Returns:
        --------
        str or DataFrame
            Formatted table
        """
        # Prepare table data
        table_data = []
        
        # Create compact header
        header = ["Parameter", "Behavior", 
                "Day Effects", "", "", "", "",
                "Night Effects", "", "", "", ""]
        subheader = ["", "", 
                    "Prob. Change", "Coef.", "OR", "p-value", "Sig.", 
                    "Prob. Change", "Coef.", "OR", "p-value", "Sig."]
        
        table_data.append(header)
        table_data.append(subheader)
        
        # Parameter labels
        param_labels = {
            '(Intercept)': 'Baseline at average predictors',
            'temp_z': 'Temperature effect (5°C increase)',
            'age_z': 'Age effect (1-year increase)',
            'weight_z': 'Weight effect (30kg increase)',
            'temp_z:age_z': 'Temperature × Age interaction',
            'temp_z:weight_z': 'Temperature × Weight interaction',
            'age_z:weight_z': 'Age × Weight interaction'
        }
        
        # Process each parameter
        for param in param_labels:
            param_label = param_labels[param]
            
            # Add parameter row (header)
            table_data.append([
                param_label,
                "",
                "", "", "", "", "",
                "", "", "", "", ""
            ])
            
            # Add state-specific rows for this parameter
            for state in self.config.analysis.hmm.states:
                # Get day results for this state and parameter
                day_result = None
                if 'day' in results and state in results['day']:
                    state_results = results['day'][state]
                    if 'fit_result' in state_results and param in state_results['fit_result'].index:
                        day_result = state_results['fit_result'].loc[param]
                
                # Get night results for this state and parameter
                night_result = None
                if 'night' in results and state in results['night']:
                    state_results = results['night'][state]
                    if 'fit_result' in state_results and param in state_results['fit_result'].index:
                        night_result = state_results['fit_result'].loc[param]
                
                # Format day values
                if day_result is not None:
                    if param == '(Intercept)':
                        # For baseline, show actual probability
                        day_prob = day_result['Prob']
                        day_prob_text = f"{day_prob*100:.1f}%"
                    else:
                        # For effects, show change in probability
                        day_prob_effect = day_result['Prob'] - 0.5
                        day_direction = "+" if day_prob_effect > 0 else ""
                        day_prob_text = f"{day_direction}{day_prob_effect*100:.1f}%"
                    
                    # Format coefficient based on parameter type
                    if param in ['temp_z', 'age_z', 'weight_z']:
                        # Back-transform standardized coefficients
                        var_name = param.replace('_z', '')
                        std = results.get('standardization_info', {}).get('stds', {}).get(var_name, 1.0)
                        
                        # Scale factors for interpretable units
                        scale_factor = 5.0 if var_name == 'temp' else (1.0 if var_name == 'age' else 30.0)
                        day_orig_coef = day_result['Estimate'] * (scale_factor / std)
                    else:
                        # Keep other coefficients as is
                        day_orig_coef = day_result['Estimate']
                    
                    day_odds_ratio = day_result['OR']
                    
                    # Format p-value
                    day_p_val = day_result['P-val']
                    if day_p_val < 0.0001:
                        day_p_val_text = "<0.0001"
                    else:
                        day_p_val_text = f"{day_p_val:.4f}"
                    
                    day_sig = day_result['Sig']
                else:
                    # No results available
                    day_prob_text = "N/A"
                    day_orig_coef = "N/A"
                    day_odds_ratio = "N/A"
                    day_p_val_text = "N/A"
                    day_sig = ""
                
                # Format night values
                if night_result is not None:
                    if param == '(Intercept)':
                        # For baseline, show actual probability
                        night_prob = night_result['Prob']
                        night_prob_text = f"{night_prob*100:.1f}%"
                    else:
                        # For effects, show change in probability
                        night_prob_effect = night_result['Prob'] - 0.5
                        night_direction = "+" if night_prob_effect > 0 else ""
                        night_prob_text = f"{night_direction}{night_prob_effect*100:.1f}%"
                    
                    # Format coefficient based on parameter type
                    if param in ['temp_z', 'age_z', 'weight_z']:
                        # Back-transform standardized coefficients
                        var_name = param.replace('_z', '')
                        std = results.get('standardization_info', {}).get('stds', {}).get(var_name, 1.0)
                        
                        # Scale factors for interpretable units
                        scale_factor = 5.0 if var_name == 'temp' else (1.0 if var_name == 'age' else 30.0)
                        night_orig_coef = night_result['Estimate'] * (scale_factor / std)
                    else:
                        # Keep other coefficients as is
                        night_orig_coef = night_result['Estimate']
                    
                    night_odds_ratio = night_result['OR']
                    
                    # Format p-value
                    night_p_val = night_result['P-val']
                    if night_p_val < 0.0001:
                        night_p_val_text = "<0.0001"
                    else:
                        night_p_val_text = f"{night_p_val:.4f}"
                    
                    night_sig = night_result['Sig']
                else:
                    # No results available
                    night_prob_text = "N/A"
                    night_orig_coef = "N/A"
                    night_odds_ratio = "N/A"
                    night_p_val_text = "N/A"
                    night_sig = ""
                
                # Format numeric values properly
                if isinstance(day_orig_coef, (int, float)):
                    day_orig_coef = f"{day_orig_coef:.4f}"
                if isinstance(day_odds_ratio, (int, float)):
                    day_odds_ratio = f"{day_odds_ratio:.4f}"
                if isinstance(night_orig_coef, (int, float)):
                    night_orig_coef = f"{night_orig_coef:.4f}"
                if isinstance(night_odds_ratio, (int, float)):
                    night_odds_ratio = f"{night_odds_ratio:.4f}"
                
                # Add row with combined day/night data
                table_data.append([
                    "",
                    state.capitalize(),
                    day_prob_text, day_orig_coef, day_odds_ratio, day_p_val_text, day_sig,
                    night_prob_text, night_orig_coef, night_odds_ratio, night_p_val_text, night_sig
                ])
        
        # Add Model Fit Statistics section - make sure it only appears once
        table_data.append([
            "Model Fit Statistics",
            "",
            "", "", "", "", "",
            "", "", "", "", ""
        ])
        
        # Add model fit statistics for each behavior state
        for state in self.config.analysis.hmm.states:
            # Get AIC/BIC values for day model
            day_aic = "N/A"
            day_bic = "N/A"
            if 'day' in results and state in results['day']:
                model = results['day'][state].get('model')
                if model is not None and hasattr(model, 'AIC'):
                    day_aic = f"AIC: {model.AIC:.1f}"
                    day_bic = f"BIC: {model.BIC:.1f}"
            
            # Get AIC/BIC values for night model
            night_aic = "N/A"
            night_bic = "N/A"
            if 'night' in results and state in results['night']:
                model = results['night'][state].get('model')
                if model is not None and hasattr(model, 'AIC'):
                    night_aic = f"AIC: {model.AIC:.1f}"
                    night_bic = f"BIC: {model.BIC:.1f}"
            
            # Add model fit row
            table_data.append([
                "",
                state.capitalize(),
                day_aic, day_bic, "", "", "",
                night_aic, night_bic, "", "", ""
            ])
        
        # Format and return table
        if format_type == "unicode":
            from tabulate import tabulate
            table_str = tabulate(table_data, headers="firstrow", tablefmt="fancy_grid")
            
            # Add note for clarity
            note = (
                "\nNotes:\n"
                "• Prob. Change: Probability change for the given parameter change\n"
                "• Coef.: Original coefficient (back-transformed for interpretability)\n"
                "• OR: Odds Ratio\n"
                "• Sig.: Statistical significance (* p<0.05, ** p<0.01, *** p<0.001, . p<0.1)"
            )
            return table_str + note
        else:
            # Create DataFrame for Excel
            df = pd.DataFrame(table_data[2:], columns=table_data[0])  # Skip header row for Excel
            df.columns = ["Parameter", "Behavior", 
                        "Day Prob Change", "Day Coef", "Day OR", "Day p-value", "Day Sig", 
                        "Night Prob Change", "Night Coef", "Night OR", "Night p-value", "Night Sig"]
            return df


    def _export_table_to_excel_compact(self, results, filename=None):
        """
        Export the compact comparison table to Excel
        
        Parameters:
        -----------
        results : dict
            Dictionary of model results
        filename : str, optional
            Path to save Excel file, defaults to temp_behavior_table.xlsx in visuals folder
        """
        # Get the table as a DataFrame
        df = self._create_comparison_table_compact(results, format_type="excel")

        # Set default filename if not provided
        if filename is None:
            filename = os.path.join(self.config.visuals.visuals_root_path, 'temp_behavior_table.xlsx')
        
        # Create a new Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Effects"
        
        # Add main header spanning both day and night
        ws.merge_cells('A1:L1')
        ws['A1'] = "Temperature Effects on Cow Behavior - Day vs. Night Comparison"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add day/night subheaders
        ws.merge_cells('C2:G2')
        ws['C2'] = "Day Effects"
        ws['C2'].font = Font(bold=True, color="FFFFFF")
        ws['C2'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['C2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('H2:L2')
        ws['H2'] = "Night Effects"
        ws['H2'].font = Font(bold=True, color="FFFFFF")
        ws['H2'].fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        ws['H2'].alignment = Alignment(horizontal='center')
        
        # Add column headers for day and night sections
        for col, header in zip(['C', 'D', 'E', 'F', 'G'], ['Prob Change', 'Coef', 'OR', 'p-value', 'Sig']):
            ws[f'{col}3'] = header
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            ws[f'{col}3'].alignment = Alignment(horizontal='center')
        
        for col, header in zip(['H', 'I', 'J', 'K', 'L'], ['Prob Change', 'Coef', 'OR', 'p-value', 'Sig']):
            ws[f'{col}3'] = header
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
            ws[f'{col}3'].alignment = Alignment(horizontal='center')
        
        # Add parameter and behavior headers
        ws['A3'] = "Parameter"
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws['B3'] = "Behavior"
        ws['B3'].font = Font(bold=True)
        ws['B3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        ws['B3'].alignment = Alignment(horizontal='center')
        
        # Write the DataFrame data starting from row 4
        row_idx = 4
        for index, row in df.iterrows():
            for col_idx, value in enumerate(row.values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
            row_idx += 1
        
        # Define styles for different row types
        param_font = Font(bold=True)
        param_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Apply styles to all data rows
        for row in range(4, row_idx):
            # Apply borders to all cells in this row
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # Check if this is a parameter header row or model statistics header
            if ws.cell(row=row, column=1).value in [
                'Baseline at average predictors', 
                'Temperature effect (5°C increase)',
                'Age effect (1-year increase)',
                'Weight effect (30kg increase)',
                'Temperature × Age interaction',
                'Temperature × Weight interaction',
                'Age × Weight interaction',
                'Model Fit Statistics'
            ]:
                # Format parameter header row
                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col)
                    cell.font = param_font
                    cell.fill = param_fill
            
            # Apply alternating colors to behavior rows - make sure we only apply to actual behavior names
            elif ws.cell(row=row, column=2).value in ['Grazing', 'Resting', 'Traveling']:
                if row % 2 == 0:  # Alternate rows
                    for col in range(1, 13):
                        ws.cell(row=row, column=col).fill = light_fill
                
        # Auto-adjust column widths - fix for MergedCell error
        for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            max_length = 0
            for cell in ws[column_letter]:
                if not isinstance(cell, (ws.merged_cells.__class__, type(None))):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fixed width overrides for certain columns
        ws.column_dimensions['A'].width = 28  # Parameter names
        ws.column_dimensions['B'].width = 12  # Behavior names
        
        # Ensure day and night columns have matching widths
        for day_col, night_col in [('C', 'H'), ('D', 'I'), ('E', 'J'), ('F', 'K'), ('G', 'L')]:
            width = max(ws.column_dimensions[day_col].width, ws.column_dimensions[night_col].width)
            ws.column_dimensions[day_col].width = width
            ws.column_dimensions[night_col].width = width
        
        # Add explanatory notes at the bottom
        note_row = row_idx + 1
        ws.merge_cells(f'A{note_row}:L{note_row}')
        ws[f'A{note_row}'] = "Notes:"
        ws[f'A{note_row}'].font = Font(bold=True)
        
        notes = [
            "• Prob Change: Absolute change in behavior probability for the given parameter change",
            "• Coefficient values are back-transformed to original units where applicable",
            "• Temperature effects are per 5°C increase; Age effects are per year; Weight effects are per 30kg",
            "• Significance: * p<0.05, ** p<0.01, *** p<0.001, . p<0.1"
        ]
        
        for i, note in enumerate(notes):
            note_cell = f'A{note_row + i + 1}'
            ws.merge_cells(f'{note_cell}:L{note_cell}')
            ws[note_cell] = note
        
        # Save the workbook
        wb.save(filename)
        print(f"Exported compact comparison table to {filename}")




    def analyze_binomial_glmm(self, df):
        """
        Main method to analyze the relationship between temperature and behavior
        using binomial GLMM with day/night as a predictor
        """
        df = self._add_cols(df)

        if len(df) > 50000:
            samp = 1
            print(f"Using a {100*samp}% random sample for faster processing ({len(df)} -> {int(len(df)*samp)} rows)")
            df = df.sample(frac=samp, random_state=42)

        results = {}
        df_day = df[df['is_daylight']].copy()
        df_night = df[~df['is_daylight']].copy()


        if self.config.visuals.temperature_graph.daynight in ["day", 'both']:
            results['day'] = self.analyze_temperature_behavior_for_period(df=df_day,  period='day', results=results)
        # results = self.analyze_temperature_behavior_glmm_by_period_MULTITHREAD(df=df, results=results)

        if self.config.visuals.temperature_graph.daynight in ["night", 'both']:
            results['night'] = self.analyze_temperature_behavior_for_period(df=df_night, period='night', results=results)

        
        # results = self.analyze_temperature_behavior_glmm(df=df, period='night', results=results)

        # results = self.analyze_temperature_behavior_glmm(df=df, results=results)

        # Create and print an interpretable table
        # table_unicode = self._create_interpretable_table(results, format_type="unicode")

        # print(results)
        table_unicode = self._create_comparison_table_compact(results, format_type="unicode")
        print("\nINTERPRETABLE MODEL RESULTS:")
        print(table_unicode)
        
        # Export table to Excel if needed
        if self.config.visuals.temperature_graph.export_excel:
            self._export_table_to_excel_compact(results)
            print("Results exported to Excel successfully.")

        return 0
