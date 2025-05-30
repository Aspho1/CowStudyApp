from typing import List
import pandas as pd
import numpy as np
# import statsmodels.api as sm
# import statsmodels.formula.api as smf
from pymer4.models import Lmer
# from statsmodels.genmod.bayes_mixed_glm import BinomialBayesMixedGLM
# import matplotlib.pyplot as plt
# import seaborn as sns
import os
# from scipy import stats
import ephem
# from datetime import date, datetime


# Import necessary libraries for parallelization
import concurrent.futures
import multiprocessing

from cowstudyapp.utils import from_posix_col
from cowstudyapp.config import ConfigManager


import sys

# For Windows systems, try to increase memory allocation
if sys.platform == 'win32':
    try:
        import ctypes
        # Allow Python to use more memory (Windows specific)
        kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
        kernel32.SetProcessWorkingSetSize(-1, -1, -1)
        print("Memory limits increased for Windows")
    except Exception as e:
        print(f"Could not optimize Windows memory: {e}")

# Configure threading for numpy/scipy
try:
    import numpy as np
    np.seterr(all='ignore')  # Ignore numerical warnings
    os.environ["OMP_NUM_THREADS"] = "4"  # Control OpenMP threads
    os.environ["MKL_NUM_THREADS"] = "4"  # Control MKL threads
    print("NumPy thread settings optimized")
except Exception as e:
    print(f"Could not configure NumPy threading: {e}")


# Set up R for better performance
def optimize_r_performance():
    try:
        import rpy2.robjects as ro
        # Use multiple cores for BLAS operations
        ro.r('library(parallel)')
        cores = max(1, multiprocessing.cpu_count() - 1)
        ro.r(f'options(mc.cores = {cores})')
        
        # Set BLAS/LAPACK threading
        ro.r('Sys.setenv(OMP_NUM_THREADS = 6)')  # Adjust based on your system
        
        print("R performance optimizations applied")
    except Exception as e:
        print(f"Could not optimize R performance: {e}")

# Call this function at the beginning of your script
optimize_r_performance()

class GrazingVersusTemperatureBayes:
    def __init__(self, config: ConfigManager):
        self.config = config
        # Set up Norris, MT observer
        self.observer = ephem.Observer()
        self.observer.lat = '45.575'
        self.observer.lon = '-111.625'
        self.observer.elevation = 1715
        self.observer.pressure = 0
        self.observer.horizon = '-0:34'
        # self._model_counter = []  # Track model numbers
        self.buffer = 1.5


    def _add_cols(self, df):
        df = self._add_suntime_cols(df)
        df = self._add_temp_col(df)
        df = self._add_meta_info(df)
        return df

    def _add_meta_info(self,df):

        meta_info_df = pd.read_excel(self.config.io.cow_info_path)
        meta_info_df["age"] = 2022 - meta_info_df['year_b']
        meta_info_df.set_index('collar_id',inplace=True)
        meta_info_df['BW_preg'] = meta_info_df['BW_preg']*0.4535924 # lbs to kg
        age_dict = meta_info_df['age'].to_dict()
        weight_dict = meta_info_df['BW_preg'].to_dict()

        df['age'] = df["ID"].map(age_dict)
        df['weight'] = df["ID"].map(weight_dict)

        df = df[df['weight'] > 0]
        return df


    def _filter_weigh_days(self, df:pd.DataFrame) -> pd.DataFrame:
        if self.config.visuals.heatmap.filter_weigh_days:
            weigh_days = pd.to_datetime(self.config.visuals.heatmap.weigh_days)
            if weigh_days is None:
                raise ValueError("Weigh Days must be defined in the config.")
            
            df = df[~df["date"].isin([wd.date() for wd in weigh_days])]

        unique_dates = df['date'].unique()
        print(f"Keeping {len(unique_dates)} unique dates after filtering weigh days...")      
        return df



    def _add_suntime_cols(self,df) -> pd.DataFrame:
        """Calculate sunrise/sunset times for all rows"""

        print("Calculating sunrise and sunset times...")
        
        # Make sure we have dates
        if 'mt' not in df.columns:
            df['mt'] = from_posix_col(df['posix_time'], self.config.analysis.timezone)
        
        # Extract date component
        df['date'] = df['mt'].dt.date

        df = self._filter_weigh_days(df)
        
        # Get unique dates
        unique_dates = df['date'].unique()
        print(f"Calculating sun times for {len(unique_dates)} unique dates...")
        

        # Calculate sunrise/sunset for each unique date
        sun_times = {}
        for date in unique_dates:
            # Calculate sunrise time
            self.observer.date = date.strftime('%Y/%m/%d 00:00:00')
            sunrise = pd.to_datetime(str(self.observer.next_rising(ephem.Sun())))
            sunrise = sunrise.tz_localize('UTC').tz_convert(self.config.analysis.timezone)
            
            # Calculate sunset time
            self.observer.date = date.strftime('%Y/%m/%d 12:00:00')
            sunset = pd.to_datetime(str(self.observer.next_setting(ephem.Sun())))
            sunset = sunset.tz_localize('UTC').tz_convert(self.config.analysis.timezone)
            

    
            sun_times[date] = {
                'sunrise': sunrise - pd.Timedelta(hours=self.buffer),
                'sunset': sunset + pd.Timedelta(hours=self.buffer)
            }
        
        # Create dataframes for merging
        sunrise_df = pd.DataFrame([
            {'date': date, 'sunrise': times['sunrise']} 
            for date, times in sun_times.items()
        ])
        
        sunset_df = pd.DataFrame([
            {'date': date, 'sunset': times['sunset']} 
            for date, times in sun_times.items()
        ])

        # Merge sunrise times
        df = pd.merge(df, sunrise_df, on='date', how='left')
        
        # Merge sunset times
        df = pd.merge(df, sunset_df, on='date', how='left')

        df['is_daylight'] = ((df['mt'] >= df['sunrise']) & (df['mt'] <= df['sunset']))


        df['minutes_after_sunrise'] = (df['mt'] - df['sunrise']).dt.total_seconds() / 60
        df['minutes_of_daylight'] = (df['sunset'] - df['sunrise']).dt.total_seconds() / 60
        
        # Create standardized time (0-1 range from sunrise to sunset)
        df['rel_time'] = df['minutes_after_sunrise'] / df['minutes_of_daylight']



        return df


    def _add_temp_col(self, df):
        """Add temperature data by joining with temperature dataset"""
        if 'temp' not in df.columns:
            if 'temperature_gps' in df.columns:
                df.rename(columns={'temperature_gps': 'temp'}, inplace=True)
            else:
                # Read temperature data
                temp_df = pd.read_csv(self.config.analysis.target_dataset)
                
                # Ensure we have the right columns for joining
                required_cols = ['device_id', 'posix_time', 'temperature_gps']
                if not all(col in temp_df.columns for col in required_cols):
                    raise ValueError(f"Temperature dataset must contain columns: {required_cols}")
                
                # Perform inner join
                df = pd.merge(
                    df,
                    temp_df[['device_id', 'posix_time', 'temperature_gps']],
                    left_on=['ID', 'posix_time'],
                    right_on=['device_id', 'posix_time'],
                    how='inner'
                )
                
                # Rename the temp column
                df.rename(columns={'temperature_gps': 'temp'}, inplace=True)
                
                # Drop the extra device_id column from the join
                df.drop('device_id', axis=1, inplace=True)
        

        return df


    def _scale_param(self, orig_var:str, row:pd.Series, baseline_prob:float, stds: dict, scale_factors: dict):
        scale_factor = scale_factors[orig_var]
        std = stds[orig_var]            
        scaled_coef = row['Estimate'] * (scale_factor / std)
        
        # Calculate marginal effect
        margin_effect = baseline_prob * (1 - baseline_prob) * scaled_coef
        prob_dir = "increase" if margin_effect > 0 else "decrease"
        return margin_effect, prob_dir


    def _scale_interaction_param(self, idx: str, row: pd.Series, baseline_prob: float, stds: dict, scale_factors: dict):
        """
        Calculate scaled effects for interaction terms
        
        Parameters:
        -----------
        idx : str
            The parameter name (e.g., 'temp_z:age_z')
        row : pd.Series
            The model result row
        baseline_prob : float
            Baseline probability
        stds : dict
            Standard deviations of variables
        scale_factors : dict
            Scaling factors for interpretable units
        
        Returns:
        --------
        tuple
            (margin_effect, interpretation_string)
        """
        # Split the interaction term
        var1, var2 = idx.split(':')
        var1_name = var1.replace('_z', '')
        var2_name = var2.replace('_z', '')
        
        # Get scaling factors and standard deviations for both variables
        scale1 = scale_factors[var1_name]
        scale2 = scale_factors[var2_name]
        std1 = stds[var1_name]
        std2 = stds[var2_name]
        
        # Calculate combined scaling factor for interaction
        combined_scale = (scale1 * scale2) / (std1 * std2)
        
        # Calculate marginal effect
        scaled_coef = row['Estimate'] * combined_scale
        margin_effect = baseline_prob * (1 - baseline_prob) * scaled_coef
        
        # Create interpretation string based on the variables involved
        if 'temp' in var1_name or 'temp' in var2_name:
            temp_effect = "increases" if margin_effect > 0 else "decreases"
            if 'age' in idx:
                interpretation = f"Temperature effect {temp_effect} by {abs(margin_effect*100):.1f}% per 5°C for each year of age"
            elif 'weight' in idx:
                interpretation = f"Temperature effect {temp_effect} by {abs(margin_effect*100):.1f}% per 5°C for each 30kg of weight"
        elif 'age_z:weight_z' == idx:
            effect_dir = "increases" if margin_effect > 0 else "decreases"
            interpretation = f"Age effect {effect_dir} by {abs(margin_effect*100):.1f}% per year for each 30kg of weight"
        
        return margin_effect, interpretation


    def plot_temp_age_interaction(self, state: str, model_results, period: str):
        """
        Create an interaction plot showing how temperature effects vary with age
        
        Parameters:
        -----------
        state : str
            The behavior state (e.g., 'Grazing')
        model_results : dict
            The model results containing coefficients and standardization info
        period : str
            'day' or 'night'
        """
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        # Get coefficients
        coef = model_results['fit_result']
        intercept = coef.loc['(Intercept)', 'Estimate']
        temp_effect = coef.loc['temp_z', 'Estimate']
        age_effect = coef.loc['age_z', 'Estimate']
        interaction = coef.loc['temp_z:age_z', 'Estimate']
        
        # Get standardization info
        means = model_results['standardization_info']['means']
        stds = model_results['standardization_info']['stds']
        
        # Create temperature and age ranges for plotting
        temp_range = np.linspace(means['temp'] - 3*stds['temp'], 
                                 means['temp'] + 3*stds['temp'], 
                                100)
        
        # Select representative ages (e.g., mean ± 1 SD)
        ages = [means['age'] - stds['age'],
                means['age'],
                means['age'] + stds['age']]
        
        age_labels = [f"Young ({ages[0]:.1f} years)",
                      f"Average ({ages[1]:.1f} years)",
                      f"Old ({ages[2]:.1f} years)"]
        
        # Create plot
        plt.figure(figsize=(10, 6))
        colors = sns.color_palette("husl", n_colors=len(ages))
        
        for age, label, color in zip(ages, age_labels, colors):
            # Convert to z-scores
            temp_z = (temp_range - means['temp']) / stds['temp']
            age_z = (age - means['age']) / stds['age']
            
            # Calculate log odds
            log_odds = (intercept + 
                    temp_effect * temp_z +
                    age_effect * age_z +
                    interaction * temp_z * age_z)
            
            # Convert to probability
            prob = 1 / (1 + np.exp(-log_odds))
            
            plt.plot(temp_range, prob, label=label, color=color, linewidth=2)
        
        plt.xlabel('Temperature (°C)')
        plt.ylabel(f'Probability of {state}')
        plt.title(f'Temperature × Age Interaction for {state} ({period.capitalize()})')
        plt.legend()
        plt.grid(True, alpha=0.3)
        
        # Add vertical line for mean temperature
        plt.axvline(x=means['temp'], color='gray', linestyle='--', alpha=0.5,
                    label='Mean temperature')
        
        plt.tight_layout()
        
        # Save plot
        plot_path = os.path.join(self.config.visuals.visuals_root_path, 
                                f'temp_age_interaction_{state.lower()}_{period}.png')
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"Interaction plot saved to: {plot_path}")


    def plot_temp_age_interaction_higherOrder(self, state: str, model_results, period: str):
        """
        Create an interaction plot showing how temperature effects vary with age,
        including polynomial temperature effects if present
        """
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        # Get coefficients
        coef = model_results['fit_result']
        intercept = coef.loc['(Intercept)', 'Estimate']
        temp_effect = coef.loc['temp_z', 'Estimate']
        
        # Check if polynomial terms exist in the model
        has_quadratic = 'temp_z_sq' in coef.index
        has_cubic = 'temp_z_cub' in coef.index
        
        # Get polynomial coefficients if they exist
        temp_sq_effect = coef.loc['temp_z_sq', 'Estimate'] if has_quadratic else 0
        temp_cub_effect = coef.loc['temp_z_cub', 'Estimate'] if has_cubic else 0
        
        age_effect = coef.loc['age_z', 'Estimate']
        
        # Get interaction terms
        temp_age_interaction = coef.loc['temp_z:age_z', 'Estimate'] if 'temp_z:age_z' in coef.index else 0
        temp_sq_age_interaction = coef.loc['temp_z_sq:age_z', 'Estimate'] if 'temp_z_sq:age_z' in coef.index else 0
        
        # Get standardization info
        means = model_results['standardization_info']['means']
        stds = model_results['standardization_info']['stds']
        
        # Create temperature range covering the observed data
        temp_min = means['temp'] - 3*stds['temp']
        temp_max = means['temp'] + 3*stds['temp']
        temp_range = np.linspace(temp_min, temp_max, 100)
        
        # Select representative ages
        ages = [means['age'] - stds['age'],
                means['age'],
                means['age'] + stds['age']]
        
        age_labels = [f"Young ({ages[0]:.1f} years)",
                    f"Average ({ages[1]:.1f} years)",
                    f"Old ({ages[2]:.1f} years)"]
        
        # Create plot
        plt.figure(figsize=(10, 6))
        colors = sns.color_palette("husl", n_colors=len(ages))
        
        for age, label, color in zip(ages, age_labels, colors):
            # Convert to z-scores
            temp_z = (temp_range - means['temp']) / stds['temp']
            age_z = (age - means['age']) / stds['age']
            
            # Calculate polynomial terms
            temp_z_sq = temp_z ** 2
            temp_z_cub = temp_z ** 3
            

            # LOOSE ENDS WITH THE `rel_time_qu` TERM
            # Calculate log odds with polynomial terms
            log_odds = (intercept + 
                    temp_effect * temp_z +
                    (temp_sq_effect * temp_z_sq if has_quadratic else 0) +
                    (temp_cub_effect * temp_z_cub if has_cubic else 0) +
                    age_effect * age_z +
                    temp_age_interaction * temp_z * age_z +
                    (temp_sq_age_interaction * temp_z_sq * age_z if 'temp_z_sq:age_z' in coef.index else 0))
            
            # Convert to probability
            prob = 1 / (1 + np.exp(-log_odds))
            
            plt.plot(temp_range, prob, label=label, color=color, linewidth=2)
        
        # Add observed data distribution as a histogram along x-axis
        # ax2 = plt.gca().twinx()
        # sns.kdeplot(df['temp'], ax=ax2, color='lightgray', fill=True, alpha=0.3)
        # ax2.set_ylabel('Data Density')
        # ax2.set_ylim(bottom=0)
        
        plt.xlabel('Temperature (°C)')
        plt.ylabel(f'Probability of {state}')
        plt.title(f'Temperature × Age Interaction for {state} ({period.capitalize()})')
        plt.legend()
        plt.grid(True, alpha=0.3)
        
        # Add vertical lines for important reference points
        plt.axvline(x=means['temp'], color='gray', linestyle='--', alpha=0.5,
                    label='Mean temperature')
        
        plt.tight_layout()
        
        # Save plot
        plot_path = os.path.join(self.config.visuals.visuals_root_path, 
                                f'temp_age_interaction_{state.lower()}_{period}.png')
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"Interaction plot saved to: {plot_path}")




    def _interpret_polynomial_effects(self, state: str, period: str, coef_dict: dict, baseline_prob: float, 
                                    means: dict, stds: dict, scale_factors: dict):
        """
        Calculate and interpret polynomial effects of temperature
        
        Parameters:
        -----------
        state : str
            The behavior state being analyzed
        period : str
            'day' or 'night'
        coef_dict : dict
            Dictionary with coefficients (must contain 'temp_z' and may contain 'temp_z_sq')
        baseline_prob : float
            The baseline probability at average values
        means : dict
            Mean values of predictors
        stds : dict
            Standard deviations of predictors
        scale_factors : dict
            Scaling factors for interpretability
            
        Returns:
        --------
        dict
            Dictionary of interpretations
        """
        has_linear = 'temp_z' in coef_dict
        has_quadratic = 'temp_z_sq' in coef_dict
        
        if not has_linear and not has_quadratic:
            return {}  # No temperature effects to interpret
            
        # Get coefficients
        temp_linear_coef = coef_dict.get('temp_z', 0)
        temp_quad_coef = coef_dict.get('temp_z_sq', 0)
        
        # Calculate mean temperature and scaling factors
        mean_temp = means['temp']
        std_temp = stds['temp']
        scale = scale_factors['temp']  # Usually 5°C
        
        # Calculate reference points for interpretation
        temp_points = {
            'mean': mean_temp,
            'low': mean_temp - std_temp,
            'high': mean_temp + std_temp
        }
        
        results = {}
        
        # If both linear and quadratic terms exist, find the inflection point
        if has_linear and has_quadratic:
            # The inflection point (in z-score scale) is where the derivative equals zero
            # For quadratic function: f(x) = a*x² + b*x + c
            # The derivative is f'(x) = 2*a*x + b
            # Setting f'(x) = 0 gives us x = -b/(2*a)
            inflection_z = -temp_linear_coef / (2 * temp_quad_coef)
            inflection_temp = mean_temp + (inflection_z * std_temp)
            
            results['inflection_point'] = inflection_temp
            
            # Determine if the curve is convex (U-shaped) or concave (∩-shaped)
            is_convex = temp_quad_coef > 0
            curve_type = "U-shaped" if is_convex else "∩-shaped"
            results['curve_type'] = curve_type
            
            # Check if the inflection point is within a reasonable range of the data
            if inflection_temp > mean_temp - 2*std_temp and inflection_temp < mean_temp + 2*std_temp:
                if is_convex:
                    results['summary'] = (
                        f"Temperature has a {curve_type} effect on {state.lower()} during {period}, "
                        f"with minimum effect at {inflection_temp:.1f}°C"
                    )
                else:
                    results['summary'] = (
                        f"Temperature has a {curve_type} effect on {state.lower()} during {period}, "
                        f"with maximum effect at {inflection_temp:.1f}°C"
                    )
            else:
                # If inflection point is outside the data range, interpret as monotonic
                if (is_convex and inflection_temp < mean_temp - 2*std_temp) or \
                (not is_convex and inflection_temp > mean_temp + 2*std_temp):
                    direction = "increases" if temp_linear_coef > 0 else "decreases"
                    results['summary'] = (
                        f"Temperature effect on {state.lower()} during {period} is non-linear: "
                        f"{state.lower()} {direction} with temperature, with stronger effects at higher temperatures"
                    )
                else:
                    direction = "decreases" if temp_linear_coef > 0 else "increases"
                    results['summary'] = (
                        f"Temperature effect on {state.lower()} during {period} is non-linear: "
                        f"{state.lower()} {direction} with temperature, with stronger effects at lower temperatures"
                    )
        
        # Calculate marginal effects at different temperature points
        marginal_effects = {}
        for point_name, temp_value in temp_points.items():
            temp_z = (temp_value - mean_temp) / std_temp
            
            # For a quadratic function: f(x) = a*x² + b*x + c
            # The derivative is f'(x) = 2*a*x + b
            # This represents the slope of the function at point x
            if has_quadratic:
                derivative_z = 2 * temp_quad_coef * temp_z + temp_linear_coef
            else:
                derivative_z = temp_linear_coef
                
            # Convert derivative from z-scale to original scale
            # derivative * (delta_orig / delta_z)
            derivative_orig = derivative_z * (scale / std_temp)
            
            # Calculate marginal effect
            marginal_effect = baseline_prob * (1 - baseline_prob) * derivative_orig
            marginal_effects[point_name] = marginal_effect
        
        results['marginal_effects'] = marginal_effects
        
        # Generate detailed interpretation
        if has_linear and has_quadratic:
            effect_at_mean = marginal_effects['mean'] * 100
            effect_at_low = marginal_effects['low'] * 100
            effect_at_high = marginal_effects['high'] * 100
            
            results['interpretation'] = (
                f"• Temperature effect on {state.lower()} is non-linear ({curve_type}):\n"
                f"  - At average temp ({mean_temp:.1f}°C): {scale}°C increase causes {effect_at_mean:.1f}% change\n"
                f"  - At lower temp ({temp_points['low']:.1f}°C): {scale}°C increase causes {effect_at_low:.1f}% change\n"
                f"  - At higher temp ({temp_points['high']:.1f}°C): {scale}°C increase causes {effect_at_high:.1f}% change\n"
            )
        else:
            # Linear effect only
            effect = marginal_effects['mean'] * 100
            direction = "increase" if effect > 0 else "decrease"
            results['interpretation'] = (
                f"• Temperature effect on {state.lower()}: {scale}°C increase causes "
                f"{abs(effect):.1f}% {direction} in probability"
            )
            
        return results




    def analyze_temperature_behavior_for_period(self, df, period, results=None, polynomial_degree=1, degree_rel_time = 0):
        """
        Analyze the relationship between temperature and behavior states
        for a specific period (day or night)
        
        Parameters:
        -----------
        df : pandas DataFrame
            Data frame with behavior states and temperature data
        period : str
            'day' or 'night'
        results : dict, optional
            Results dictionary to update
            
        Returns:
        --------
        dict
            Dictionary of model results for each state in the specified period
        """
        if results is None:
            results = {}
            
        print(f"\n{'-'*80}")
        print(f"ANALYZING TEMPERATURE EFFECTS ON BEHAVIOR DURING {period.upper()}")
        print(f"{'-'*80}")
        
        # Try to optimize R memory usage
        try:
            import rpy2.robjects as ro
            ro.r('gc()')  # Run garbage collection in R
            ro.r('options(warn=-1)')  # Suppress R warnings
        except Exception as e:
            print(f"Could not configure R: {e}")

        # Print data summary
        print(f"\nData Summary for {period}:")
        print(f"Temperature range: {df['temp'].min():.1f} to {df['temp'].max():.1f}°C")
        print(f"Age range: {df['age'].min()} to {df['age'].max()} years")
        print(f"Weight range: {df['weight'].min():.1f} to {df['weight'].max():.1f} kg")
        print(f"Number of cows: {df['ID'].nunique()}")
        print(f"Total observations: {len(df)}")
        
        # Store original means and standard deviations for back-transformation
        means = {
            'temp': df['temp'].mean(),
            'age': df['age'].mean(),
            'weight': df['weight'].mean()
        }
        
        stds = {
            'temp': df['temp'].std(),
            'age': df['age'].std(),
            'weight': df['weight'].std()
        }
        
        selected_factors = ['temp_z', 'age_z', 'weight_z']
        
        # Systematically add polynomial terms
        if polynomial_degree >= 2:
            selected_factors.append('temp_z_sq')
        if polynomial_degree >= 3:
            selected_factors.append('temp_z_cub')
        
        # Process the clean DataFrame
        clean_df = pd.DataFrame()
        clean_df['ID'] = df['ID']
        
        # Add z-standardized variables for all base variables
        for factor in ['temp', 'age', 'weight']:
            z_col = f'{factor}_z'
            clean_df[z_col] = (df[factor] - means[factor]) / stds[factor]
        
        selected_interactions = None
        selected_factors = [
            'temp_z',
            'age_z',
            'weight_z'
        ]

        selected_interactions = [
            'temp_z:age_z',
            'temp_z:weight_z',
            'age_z:weight_z'
        ]
        
        # Add polynomial terms
        if polynomial_degree >= 2:
            clean_df['temp_z_sq'] = clean_df['temp_z'] ** 2
            selected_factors.append('temp_z_sq')
            if polynomial_degree >= 3:
                clean_df['temp_z_cub'] = clean_df['temp_z'] ** 3
                selected_factors.append('temp_z_cub')


        if degree_rel_time == 1:
            selected_factors.append('rel_time')
        elif degree_rel_time == 2:
            selected_factors.append('rel_time_sq')
            clean_df['rel_time_sq'] = clean_df['rel_time'] ** degree_rel_time
            
        elif degree_rel_time == 3:
            selected_factors.append('rel_time_cub')
            clean_df['rel_time_cub'] = clean_df['rel_time'] ** degree_rel_time
            
        elif degree_rel_time == 4:
            selected_factors.append('rel_time_qu')
            selected_interactions.append("temp:rel_time_qu")
            clean_df['rel_time_qu'] = clean_df['rel_time'] ** degree_rel_time
            
        


        # Create indicator variables for each state
        for state in self.config.analysis.hmm.states:
            clean_df[f'is_{state.lower()}'] = (df['predicted_state'] == state).astype(int)
        
        # Define fixed effects formula
        fixed_effects = ' + '.join(selected_factors)
        if selected_interactions is not None:
            fixed_effects += ' + ' + ' + '.join(selected_interactions)
        

        random_effects = "(1|ID)"  # Random intercepts for cow ID
        # random_effects = "(1 + temp_z|ID)"  # Random slopes for temperature by cow ID
        
        # Create formula template
        formula_template = 'is_{} ~ {} + {}'
        
        # Store standardization info for back-transformation
        results['standardization_info'] = {
            'means': means,
            'stds': stds
        }
        # Scaling factors for interpretable units
        scale_factors = {
            'temp': 5.0,  # per 5°C
            'age': 1.0,   # per year
            'weight': 30.0  # per 30kg
        }
        results['scale_factors'] = scale_factors

        # Store model specification
        results['model_specification'] = {
            'selected_factors': selected_factors,
            'selected_interactions': selected_interactions
        }
        
        # Store clean data
        results['data'] = clean_df.copy()


        # Analyze each behavior state
        state_results = {}
        for state in self.config.analysis.hmm.states:
            print(f"\n{'-'*40}")
            print(f"ANALYZING {state.upper()} STATE - {period.upper()}")
            print(f"{'-'*40}")
            
            # Create formula for this state
            formula = formula_template.format(state.lower(), fixed_effects, random_effects)
            print(f"Model formula: {formula}")
            
            try:
                # Run R garbage collection before fitting
                import gc
                gc.collect()
                
                try:
                    import rpy2.robjects as ro
                    ro.r('gc()')
                except:
                    pass
                    
                # Fit the model
                model = Lmer(formula, data=clean_df, family='binomial')
                fit_result = model.fit()
                
                # Print model results
                print(f"\nModel Results for {state} ({period}):")
                print(fit_result)
                
                # Store results
                state_results[state] = {
                    'formula': formula,
                    'model': model,
                    'fit_result': fit_result,
                    'standardization_info': results['standardization_info']
                }
                # Baseline probability
                baseline_prob = fit_result.loc['(Intercept)', 'Prob']

                if 'temp_z:age_z' in state_results[state]['fit_result'].index:
                    # self.plot_temp_age_interaction(state, state_results[state], period)
                    self.plot_temp_age_interaction_higherOrder(state, state_results[state], period)
                
                
                
                # Create a dictionary of all coefficients for easier access
                coef_dict = {idx: row['Estimate'] for idx, row in fit_result.iterrows()}
                
                # Interpret polynomial effects
                poly_interpretation = self._interpret_polynomial_effects(
                    state, period, coef_dict, baseline_prob, means, stds, scale_factors
                )
                
                # Store polynomial interpretation in results
                state_results[state]['polynomial_effects'] = poly_interpretation
                
                # Print interpretable results including polynomial effects
                print(f"\nInterpretable Results for {state} ({period}):")
                
                # Print baseline probability
                print(f"• Baseline probability of {state} during {period}: {baseline_prob:.3f} ({baseline_prob*100:.1f}%)")
                
                # If we have polynomial effects, print the interpretation
                if poly_interpretation and 'interpretation' in poly_interpretation:
                    print(poly_interpretation['interpretation'])
                    
                    # If we found an inflection point, print that too
                    if 'inflection_point' in poly_interpretation:
                        print(f"• Inflection point: {poly_interpretation['inflection_point']:.1f}°C")
                

                for idx, row in fit_result.iterrows():

                    if idx in ['(Intercept)', 'temp_z', 'temp_z_sq', 'temp_z_cub']:
                        continue                    
                        
                    # For effect parameters, calculate marginal effect at average values
                    elif idx == 'temp_z':
                        orig_var = 'temp'
                        margin_effect, prob_dir = self._scale_param(orig_var, row, baseline_prob, stds, scale_factors)
                        print(f"• Temperature effect ({period}): 5°C increase causes {abs(margin_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    
                    elif idx == 'age_z':
                        orig_var = 'age'
                        margin_effect, prob_dir = self._scale_param(orig_var, row, baseline_prob, stds, scale_factors)
                        print(f"• Age effect ({period}): 1 year increase causes {abs(margin_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")

                    elif idx == 'weight_z':
                        orig_var = 'weight'
                        margin_effect, prob_dir = self._scale_param(orig_var, row, baseline_prob, stds, scale_factors)
                        print(f"• Weight effect ({period}): 30kg increase causes {abs(margin_effect*100):.1f}% {prob_dir} in probability {row['Sig']}")
                    

                    # Handle interaction terms
                    elif ':' in idx:
                        margin_effect, interpretation = self._scale_interaction_param(
                            idx, row, baseline_prob, stds, scale_factors
                        )
                        print(f"• Interaction effect: {interpretation} {row['Sig']}")
                    
                    else:
                        # Generic handler for other effects
                        print(f"• {idx}: Unexpected parameter {row['Sig']}")                

                
                # Print model fit statistics
                if hasattr(model, 'AIC'):
                    print(f"\nModel Fit Statistics:")
                    print(f"  AIC: {model.AIC:.1f}")
                    print(f"  BIC: {model.BIC:.1f}")
                
            except Exception as e:
                print(f"Error fitting model for {state} during {period}: {str(e)}")
                import traceback
                traceback.print_exc()
                state_results[state] = {'error': str(e)}
                
            # Clean up memory
            gc.collect()
        
        # Store all results for this period
        return state_results





    def _add_row_to_table(self, param: str, row_data: List, state: str, period_result, results) -> List:
        """
        Add formatted data for a specific parameter, state and period to a table row
        with improved probability calculations for higher order terms
        
        Parameters:
        -----------
        param : str
            The parameter name (e.g., 'temp_z', 'age_z')
        row_data : list
            The current row data to add to
        state : str
            The behavior state being analyzed
        period_result : pd.Series
            The model result for this parameter
        results : dict
            Complete results dictionary with standardization info
            
        Returns:
        --------
        list
            Updated row data with formatted values
        """
        if period_result is None:
            # No results available for this parameter/state/period
            return row_data + ["N/A", "N/A", "N/A", "N/A", "N/A"]
        
        # Get the intercept value for this state's model
        intercept = None
        for period_key in results:
            if state in results[period_key] and 'fit_result' in results[period_key][state]:
                if '(Intercept)' in results[period_key][state]['fit_result'].index:
                    intercept = results[period_key][state]['fit_result'].loc['(Intercept)', 'Estimate']
                    break
        
        # Format probability change using a more conservative approach
        if param == '(Intercept)':
            # For baseline, show actual probability (converted from log-odds)
            log_odds = period_result['Estimate']
            prob = 1 / (1 + np.exp(-log_odds))
            prob_text = f"{prob*100:.1f}%"
        else:
            # For effects, calculate the marginal effect at average values
            if intercept is not None:
                # Calculate baseline probability
                baseline_prob = 1 / (1 + np.exp(-intercept))
                
                # Calculate new probability with this parameter's effect
                if param in ['temp_z', 'age_z', 'weight_z']:
                    # Back-transform standardized coefficients for realistic units
                    var_name = param.replace('_z', '')
                    std = results.get('standardization_info', {}).get('stds', {}).get(var_name, 1.0)
                    
                    # Scale factors for interpretable units
                    scale_factor = results['scale_factors'].get(var_name, 1)
                    
                    # Convert to effect for the appropriate scale unit
                    scaled_effect = period_result['Estimate'] * (scale_factor / std)
                    
                    # Calculate probability with the scaled effect 
                    # Use a more conservative approach (first-order approximation)
                    margin_effect = baseline_prob * (1 - baseline_prob) * scaled_effect
                    prob_change = margin_effect * 100  # Convert to percentage
                    
                    direction = "+" if prob_change > 0 else ""
                    prob_text = f"{direction}{prob_change:.1f}%"
                    
                elif param == 'temp_z_sq':
                    # For quadratic terms, we need to show the nature of the curve
                    # Get polynomial effects information
                    poly_effects = None
                    for period_key in results:
                        if (state in results[period_key] and 
                            'polynomial_effects' in results[period_key][state]):
                            poly_effects = results[period_key][state]['polynomial_effects']
                            break
                    
                    if poly_effects and 'curve_type' in poly_effects:
                        # Show curve type instead of a direct percentage
                        curve_type = poly_effects['curve_type']
                        
                        # If inflection point is available, show that too
                        if 'inflection_point' in poly_effects:
                            inflection = poly_effects['inflection_point']
                            prob_text = f"{curve_type} (max/min at {inflection:.1f}°C)"
                        else:
                            prob_text = curve_type
                        
                        # For debugging, can include the coefficient sign
                        # coef_sign = "+" if period_result['Estimate'] > 0 else "-"
                        # prob_text = f"{curve_type} ({coef_sign})"
                    elif period_result['Estimate'] > 0:
                        prob_text = "U-shaped"  # Positive quadratic = U shape
                    else:
                        prob_text = "∩-shaped"  # Negative quadratic = ∩ shape
                
                elif param == 'temp_z_cub':
                    # For cubic terms, describe S-curve or complex effects
                    sign = "+" if period_result['Estimate'] > 0 else "-"
                    prob_text = f"S-curve ({sign})"
                    
                elif ':' in param:
                    # For interactions, use the actual coefficient
                    margin_effect = baseline_prob * (1 - baseline_prob) * period_result['Estimate']
                    prob_change = margin_effect * 100  # Convert to percentage
                    
                    direction = "+" if prob_change > 0 else ""
                    prob_text = f"{direction}{prob_change:.1f}%"
                
                else:
                    # Generic handler for other parameters
                    margin_effect = baseline_prob * (1 - baseline_prob) * period_result['Estimate']
                    prob_change = margin_effect * 100
                    
                    direction = "+" if prob_change > 0 else ""
                    prob_text = f"{direction}{prob_change:.1f}%"
            else:
                # No intercept available, use simple conversion
                prob_effect = period_result['Prob'] - 0.5
                direction = "+" if prob_effect > 0 else ""
                prob_text = f"{direction}{prob_effect*100:.1f}%"
        
        # Format coefficient based on parameter type
        if param in ['temp_z', 'age_z', 'weight_z']:
            # Back-transform standardized coefficients
            var_name = param.replace('_z', '')
            std = results.get('standardization_info', {}).get('stds', {}).get(var_name, 1.0)
            
            # Scale factors for interpretable units
            scale_factor = results['scale_factors'].get(var_name, 1)
            orig_coef = period_result['Estimate'] * (scale_factor / std)
        else:
            # Keep other coefficients as is
            orig_coef = period_result['Estimate']
        
        # Format odds ratio
        odds_ratio = period_result['OR']
        
        # Format p-value
        p_val = period_result['P-val']
        if p_val < 0.0001:
            p_val_text = "<0.0001"
        else:
            p_val_text = f"{p_val:.4f}"
        
        # Get significance
        sig = period_result['Sig']
        
        # Format numeric values properly
        if isinstance(orig_coef, (int, float)):
            orig_coef = f"{orig_coef:.4f}"
        if isinstance(odds_ratio, (int, float)):
            odds_ratio = f"{odds_ratio:.4f}"
        
        # Add all formatted values to the row
        return row_data + [prob_text, orig_coef, odds_ratio, p_val_text, sig]


    def _create_comparison_table_compact(self, results, format_type="unicode"):
        """
        Create a more compact table comparing day vs night behavior effects
        based on configuration settings, showing only terms that exist in the model
        
        Parameters:
        -----------
        results : dict
            Dictionary with day and night results
        format_type : str
            'unicode' or 'excel'
            
        Returns:
        --------
        str or DataFrame
            Formatted table
        """
        # Prepare table data
        table_data = []
        
        # Determine which periods to include based on configuration
        periods = []
        if 'day' in results and self.config.visuals.temperature_graph.daynight in ['day', 'both']:
            periods.append('day')
        if 'night' in results and self.config.visuals.temperature_graph.daynight in ['night', 'both']:
            periods.append('night')
        if 'both' in results and self.config.visuals.temperature_graph.daynight == 'both':
            periods.append('both')
        
        # Create appropriate headers based on included periods
        header = ["Parameter", "Behavior"]
        subheader = ["", ""]
        
        for period in periods:
            period_title = f"{period.capitalize()} Effects"
            header.extend([period_title, "", "", "", ""])
            subheader.extend(["Prob. Change", "Coef.", "OR", "p-value", "Sig."])
        
        table_data.append(header)
        table_data.append(subheader)
        
        # Parameter labels for better readability
        param_labels = {
            '(Intercept)': 'Baseline at average predictors',
            'temp_z': 'Temperature effect (5°C increase)',
            'temp_z_sq': 'Temperature effect (quadratic)',
            'temp_z_cub': 'Temperature effect (cubic)',
            'age_z': 'Age effect (1-year increase)',
            'weight_z': 'Weight effect (30kg increase)',
            'temp_z:age_z': 'Temperature × Age interaction',
            'temp_z:weight_z': 'Temperature × Weight interaction',
            'age_z:weight_z': 'Age × Weight interaction'
        }
        
        # Find out which parameters actually exist in each model
        existing_params = ['(Intercept)']  # Intercept always exists
        
        for period in periods:
            if period in results:
                for state in self.config.analysis.hmm.states:
                    if state in results[period] and 'fit_result' in results[period][state]:
                        for param in results[period][state]['fit_result'].index:
                            if param not in existing_params:
                                existing_params.append(param)
        
        # Only process parameters that exist in at least one model
        for param in existing_params:
            if param in param_labels:
                param_label = param_labels[param]
                
                # Add parameter header row
                table_data.append([param_label, ""])
                
                # Add state-specific rows for this parameter
                for state in self.config.analysis.hmm.states:
                    row_data = ["", state.capitalize()]
                    
                    # Add data for each period
                    for period in periods:
                        # Get results for this state and parameter for the current period
                        period_result = None
                        if period in results and state in results[period]:
                            state_results = results[period][state]
                            if 'fit_result' in state_results and param in state_results['fit_result'].index:
                                period_result = state_results['fit_result'].loc[param]
                        
                        # Add formatted values for this period
                        row_data = self._add_row_to_table(param, row_data, state, period_result, results)
                    
                    # Add completed row to table
                    table_data.append(row_data)
        
        # Add Model Fit Statistics section
        table_data.append(["Model Fit Statistics", ""])
        
        # Add model fit statistics for each behavior state
        for state in self.config.analysis.hmm.states:
            row_data = ["", state.capitalize()]
            
            # Add statistics for each period
            for period in periods:
                aic = "N/A"
                bic = "N/A"
                
                if period in results and state in results[period]:
                    model = results[period][state].get('model')
                    if model is not None and hasattr(model, 'AIC'):
                        aic = f"AIC: {model.AIC:.1f}"
                        bic = f"BIC: {model.BIC:.1f}"
                
                row_data.extend([aic, bic, "", "", ""])
            
            # Add completed stats row
            table_data.append(row_data)
        
        # Format and return table based on requested format
        if format_type == "unicode":
            from tabulate import tabulate
            table_str = tabulate(table_data, headers="firstrow", tablefmt="fancy_grid")
            
            # Add explanatory note
            note = (
                "\nNotes:\n"
                "• Prob. Change: Probability change for the given parameter change\n"
                "• Coef.: Original coefficient (back-transformed for interpretability)\n"
                "• OR: Odds Ratio\n"
                "• Sig.: Statistical significance (* p<0.05, ** p<0.01, *** p<0.001, . p<0.1)\n"
                "• Temperature effects are per 5°C increase; Age effects are per year; Weight effects are per 30kg"
            )
            return table_str + note
        else:
            # Create DataFrame for Excel
            # First create appropriate column names based on included periods
            columns = ["Parameter", "Behavior"]
            for period in periods:
                period_cap = period.capitalize()
                columns.extend([
                    f"{period_cap} Prob Change", 
                    f"{period_cap} Coef", 
                    f"{period_cap} OR", 
                    f"{period_cap} p-value", 
                    f"{period_cap} Sig"
                ])
            
            # Create DataFrame without the header row (it's handled in Excel export)
            df = pd.DataFrame(table_data[2:])  # Skip the header and subheader
            df.columns = columns
            return df


    def _export_table_to_excel_compact(self, results, filename=None):
        """
        Export the compact comparison table to Excel, respecting configuration settings
        
        Parameters:
        -----------
        results : dict
            Dictionary of model results
        filename : str, optional
            Path to save Excel file, defaults to temp_behavior_table.xlsx in visuals folder
        """
        # Get the table as a DataFrame
        df = self._create_comparison_table_compact(results, format_type="excel")

        # Determine which periods are included
        periods = []
        if 'day' in results and self.config.visuals.temperature_graph.daynight in ['day', 'both']:
            periods.append('day')
        if 'night' in results and self.config.visuals.temperature_graph.daynight in ['night', 'both']:
            periods.append('night')

        # Set default filename if not provided
        if filename is None:
            filename = os.path.join(self.config.visuals.visuals_root_path, 'temp_behavior_table.xlsx')
        
        # Create a new Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Effects"
        
        # Define column letters based on included periods
        period_columns = {}
        if len(periods) == 1:
            # Just one period
            period_columns[periods[0]] = ['C', 'D', 'E', 'F', 'G']
        elif len(periods) == 2:
            # Both day and night
            period_columns['day'] = ['C', 'D', 'E', 'F', 'G']
            period_columns['night'] = ['H', 'I', 'J', 'K', 'L']
        
        # Calculate the last column letter
        last_col = chr(ord('B') + (5 * len(periods)))
        
        # Add main header spanning all columns
        title = "Temperature Effects on Cow Behavior"
        if len(periods) == 2:
            title += " - Day vs. Night Comparison"
        elif len(periods) == 1:
            title += f" - {periods[0].capitalize()} Analysis"
        
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add period subheaders
        period_colors = {'day': "4F81BD", 'night': "9BBB59"}
        subheader_colors = {'day': "B8CCE4", 'night': "D8E4BC"}
        
        for period in periods:
            col_start = period_columns[period][0]
            col_end = period_columns[period][-1]
            ws.merge_cells(f'{col_start}2:{col_end}2')
            ws[f'{col_start}2'] = f"{period.capitalize()} Effects"
            ws[f'{col_start}2'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col_start}2'].fill = PatternFill(start_color=period_colors[period], end_color=period_colors[period], fill_type="solid")
            ws[f'{col_start}2'].alignment = Alignment(horizontal='center')
            
            # Add column headers for this period
            for col, header in zip(period_columns[period], ['Prob Change', 'Coef', 'OR', 'p-value', 'Sig']):
                ws[f'{col}3'] = header
                ws[f'{col}3'].font = Font(bold=True)
                ws[f'{col}3'].fill = PatternFill(start_color=subheader_colors[period], end_color=subheader_colors[period], fill_type="solid")
                ws[f'{col}3'].alignment = Alignment(horizontal='center')
        
        # Add parameter and behavior headers
        ws['A3'] = "Parameter"
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws['B3'] = "Behavior"
        ws['B3'].font = Font(bold=True)
        ws['B3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        ws['B3'].alignment = Alignment(horizontal='center')
        
        # Write the DataFrame data starting from row 4
        row_idx = 4
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row.values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
            row_idx += 1
        
        # Define styles for different row types
        param_font = Font(bold=True)
        param_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Number of columns to format
        last_col_idx = 2 + (5 * len(periods))
        
        # Apply styles to all data rows
        for row in range(4, row_idx):
            # Apply borders to all cells in this row
            for col in range(1, last_col_idx + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # Check if this is a parameter header row or model statistics header
            if ws.cell(row=row, column=1).value in [
                'Baseline at average predictors', 
                'Temperature effect (5°C increase)',
                'Age effect (1-year increase)',
                'Weight effect (30kg increase)',
                'Temperature × Age interaction',
                'Temperature × Weight interaction',
                'Age × Weight interaction',
                'Model Fit Statistics'
            ]:
                # Format parameter header row
                for col in range(1, last_col_idx + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = param_font
                    cell.fill = param_fill
            
            # Apply alternating colors to behavior rows
            elif ws.cell(row=row, column=2).value in ['Grazing', 'Resting', 'Traveling']:
                if row % 2 == 0:  # Alternate rows
                    for col in range(1, last_col_idx + 1):
                        ws.cell(row=row, column=col).fill = light_fill
        
        # Auto-adjust column widths
        all_columns = ['A', 'B']
        for period in periods:
            all_columns.extend(period_columns[period])
        
        for column_letter in all_columns:
            max_length = 0
            for cell in ws[column_letter]:
                if not isinstance(cell, (ws.merged_cells.__class__, type(None))):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Fixed width overrides for certain columns
        ws.column_dimensions['A'].width = 28  # Parameter names
        ws.column_dimensions['B'].width = 12  # Behavior names
        
        # Ensure columns for each period have matching widths
        for period in periods:
            cols = period_columns[period]
            max_width = max(ws.column_dimensions[col].width for col in cols)
            for col in cols:
                ws.column_dimensions[col].width = max_width
        
        # Add explanatory notes at the bottom
        note_row = row_idx + 1
        ws.merge_cells(f'A{note_row}:{last_col}{note_row}')
        ws[f'A{note_row}'] = "Notes:"
        ws[f'A{note_row}'].font = Font(bold=True)
        
        notes = [
            "• Prob Change: Absolute change in behavior probability for the given parameter change",
            "• Coefficient values are back-transformed to original units where applicable",
            "• Temperature effects are per 5°C increase; Age effects are per year; Weight effects are per 30kg",
            "• Significance: * p<0.05, ** p<0.01, *** p<0.001, . p<0.1"
        ]
        
        for i, note in enumerate(notes):
            note_cell = f'A{note_row + i + 1}'
            ws.merge_cells(f'{note_cell}:{last_col}{note_cell}')
            ws[note_cell] = note
        
        # Save the workbook
        wb.save(filename)
        print(f"Exported comparison table to {filename}")


    def _plot_model_comparison(self, linear_results, quadratic_results, period):
        """
        Create plots comparing linear and quadratic model predictions
        
        Parameters:
        -----------
        linear_results : dict
            Results from linear model
        quadratic_results : dict
            Results from quadratic model
        period : str
            'day' or 'night'
        """
        import matplotlib.pyplot as plt
        import seaborn as sns
        from matplotlib.gridspec import GridSpec
        
        # Create a figure with subplots for each behavior state
        states = self.config.analysis.hmm.states
        fig = plt.figure(figsize=(15, 5 * len(states)))
        gs = GridSpec(len(states), 2, figure=fig, width_ratios=[2, 1])
        
        # Get standardization info (should be the same for both models)
        means = linear_results[states[0]]['standardization_info']['means']
        stds = linear_results[states[0]]['standardization_info']['stds']
        
        # Create temperature range for predictions
        temp_min = means['temp'] - 2.5 * stds['temp']
        temp_max = means['temp'] + 2.5 * stds['temp']
        temp_range = np.linspace(temp_min, temp_max, 100)
        temp_range_z = (temp_range - means['temp']) / stds['temp']
        
        # Get histogram data for actual temperature distribution
        # Find any data source in the results that contains the original data
        data_source = None
        for state in states:
            if 'data' in quadratic_results[state]:
                data_source = quadratic_results[state]['data']
                break
        
        # Loop through each behavior state
        for i, state in enumerate(states):
            # Create main plot for probability curves
            ax1 = fig.add_subplot(gs[i, 0])
            
            # Get model coefficients
            linear_coefs = linear_results[state]['fit_result']
            quad_coefs = quadratic_results[state]['fit_result']
            
            # Basic model parameters (intercept and linear term)
            linear_intercept = linear_coefs.loc['(Intercept)', 'Estimate']
            linear_temp = linear_coefs.loc['temp_z', 'Estimate']
            
            quad_intercept = quad_coefs.loc['(Intercept)', 'Estimate']
            quad_temp = quad_coefs.loc['temp_z', 'Estimate']
            
            # Safely check if temp_z_sq exists in the DataFrame
            if 'temp_z_sq' in quad_coefs.index:
                quad_temp_sq = quad_coefs.loc['temp_z_sq', 'Estimate']
            else:
                quad_temp_sq = 0
            
            # Calculate predictions for both models
            linear_logodds = linear_intercept + linear_temp * temp_range_z
            linear_probs = 1 / (1 + np.exp(-linear_logodds))
            
            quad_logodds = quad_intercept + quad_temp * temp_range_z + quad_temp_sq * temp_range_z**2
            quad_probs = 1 / (1 + np.exp(-quad_logodds))
            
            # Plot both model predictions
            ax1.plot(temp_range, linear_probs, 'b-', linewidth=2, label='Linear model')
            ax1.plot(temp_range, quad_probs, 'r-', linewidth=2, label='Quadratic model')
            
            # Get AIC values for the title
            linear_aic = linear_results[state]['model'].AIC
            quad_aic = quadratic_results[state]['model'].AIC
            aic_diff = linear_aic - quad_aic
            
            # Add vertical line for mean temperature
            ax1.axvline(x=means['temp'], color='gray', linestyle='--', alpha=0.5)
            
            # Add annotations for the mean and ±1 SD temperatures
            for temp, label in [
                (means['temp'] - stds['temp'], f"μ-σ = {means['temp'] - stds['temp']:.1f}°C"),
                (means['temp'], f"μ = {means['temp']:.1f}°C"),
                (means['temp'] + stds['temp'], f"μ+σ = {means['temp'] + stds['temp']:.1f}°C")
            ]:
                ax1.axvline(x=temp, color='gray', linestyle=':', alpha=0.3)
                ax1.text(temp, 0.02, label, rotation=90, va='bottom', ha='right', fontsize=9, alpha=0.7)
            
            # Plot styling
            ax1.set_xlabel('Temperature (°C)')
            ax1.set_ylabel(f'Probability of {state}')
            ax1.set_title(f'{state} ~ Temperature ({period.capitalize()}) | AIC diff: {aic_diff:.1f}')
            ax1.grid(True, alpha=0.3)
            ax1.legend(loc='best')
            
            # Add temperature data distribution with a secondary y-axis
            if data_source is not None:
                # Add temperature histogram
                ax_top = ax1.twiny()
                ax_top.set_xticklabels([])  # Hide top x-axis labels
                
                # Try to get temperature data for the histogram
                try:
                    # Get raw temp data from z-scores
                    temps = data_source['temp_z'] * stds['temp'] + means['temp']
                    
                    # Plot density curve
                    sns.kdeplot(temps, ax=ax_top, color='green', alpha=0.3, fill=True, 
                            label='Temperature distribution')
                    
                    # Adjust top axis limits to match bottom axis
                    ax_top.set_xlim(ax1.get_xlim())
                except Exception as e:
                    print(f"Could not add temperature distribution: {e}")
            
            # Add marginal effects plot
            ax2 = fig.add_subplot(gs[i, 1])
            
            # Calculate marginal effects for both models across temperature range
            baseline_linear_prob = 1 / (1 + np.exp(-linear_intercept))
            baseline_quad_prob = 1 / (1 + np.exp(-quad_intercept))
            
            # Linear model: constant marginal effect
            linear_marginal = baseline_linear_prob * (1 - baseline_linear_prob) * linear_temp * 5 / stds['temp']
            linear_margins = np.ones_like(temp_range) * linear_marginal * 100  # Convert to percentage
            
            # Quadratic model: varying marginal effect
            quad_margins = []
            for t, t_z in zip(temp_range, temp_range_z):
                # For quadratic, derivative is: b1 + 2*b2*x
                derivative_z = quad_temp + 2 * quad_temp_sq * t_z
                # Convert to original scale
                derivative_orig = derivative_z * 5 / stds['temp']  # For 5°C increase
                # Calculate marginal effect
                local_prob = 1 / (1 + np.exp(-(quad_intercept + quad_temp * t_z + quad_temp_sq * t_z**2)))
                margin = local_prob * (1 - local_prob) * derivative_orig * 100  # As percentage
                quad_margins.append(margin)
            
            # Plot marginal effects
            ax2.plot(temp_range, linear_margins, 'b-', linewidth=2, label='Linear')
            ax2.plot(temp_range, quad_margins, 'r-', linewidth=2, label='Quadratic')
            ax2.axhline(y=0, color='k', linestyle='-', alpha=0.3)
            
            # Plot styling
            ax2.set_xlabel('Temperature (°C)')
            ax2.set_ylabel('% change per 5°C')
            ax2.set_title('Marginal Effects')
            ax2.grid(True, alpha=0.3)
            
            # Add vertical line for where effect is zero (if it exists)
            if min(quad_margins) < 0 and max(quad_margins) > 0:
                # Find where the quadratic effect crosses zero
                for i in range(1, len(temp_range)):
                    if (quad_margins[i-1] < 0 and quad_margins[i] > 0) or \
                    (quad_margins[i-1] > 0 and quad_margins[i] < 0):
                        zero_temp = (temp_range[i-1] + temp_range[i]) / 2
                        ax2.axvline(x=zero_temp, color='r', linestyle='--', alpha=0.7)
                        ax2.text(zero_temp, ax2.get_ylim()[0], f"Effect\nreversal\n{zero_temp:.1f}°C", 
                            ha='center', va='bottom', fontsize=8, bbox=dict(facecolor='white', alpha=0.5))
                        break
        
        # Adjust layout and save
        plt.tight_layout()
        plot_path = os.path.join(self.config.visuals.visuals_root_path, 
                            f'model_comparison_{period}.png')
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"Model comparison plot saved to: {plot_path}")




    def compare_polynomial_models(self, df, period):
        """Compare linear vs quadratic temperature models"""
        results = {}
        
        # First fit linear model (base model)
        print("Fitting LINEAR temperature model...")
        linear_results = self.analyze_temperature_behavior_for_period(
            df=df, 
            period=period, 
            polynomial_degree=1  # Linear only
        )
        
        # Then fit quadratic model
        print("Fitting QUADRATIC temperature model...")
        quadratic_results = self.analyze_temperature_behavior_for_period(
            df=df,
            period=period, 
            polynomial_degree=2  # Add quadratic term
        )
        
        # Compare model fits
        print("\nMODEL COMPARISON: Linear vs. Quadratic")
        print("-" * 50)
        
        for state in self.config.analysis.hmm.states:
            print(f"\n{state} behavior:")
            
            # Get AIC/BIC for each model
            linear_aic = linear_results[state]['model'].AIC
            linear_bic = linear_results[state]['model'].BIC
            quad_aic = quadratic_results[state]['model'].AIC
            quad_bic = quadratic_results[state]['model'].BIC
            
            # Calculate improvements
            aic_diff = linear_aic - quad_aic
            bic_diff = linear_bic - quad_bic
            
            print(f"  AIC: Linear={linear_aic:.1f}, Quadratic={quad_aic:.1f}, Improvement={aic_diff:.1f}")
            print(f"  BIC: Linear={linear_bic:.1f}, Quadratic={quad_bic:.1f}, Improvement={bic_diff:.1f}")
            
            # Get quadratic term coefficient and significance
            quad_term = quadratic_results[state]['fit_result'].get('temp_z_sq', None)
            if quad_term is not None:
                print(f"  Quadratic term: coef={quad_term['Estimate']:.4f}, p={quad_term['P-val']:.4f} {quad_term['Sig']}")
            
            # Suggest which model to use
            if aic_diff > 2 and quad_term is not None and quad_term['P-val'] < 0.05:
                recommendation = "INCLUDE quadratic term"
            else:
                recommendation = "KEEP simpler linear model"
            
            print(f"  Recommendation: {recommendation}")
        
        # Create plots to visualize differences
        self._plot_model_comparison(linear_results, quadratic_results, period)
        
        return {"linear": linear_results, "quadratic": quadratic_results}


    def analyze_binomial_glmm(self, df):
        """
        Main method to analyze the relationship between temperature and behavior
        using binomial GLMM with day/night as a predictor
        """
        df = self._add_cols(df)

        if len(df) > 50000:
            samp = 1
            print(f"Using a {100*samp}% random sample for faster processing ({len(df)} -> {int(len(df)*samp)} rows)")
            df = df.sample(frac=samp, random_state=42)

        results = {}
        df_day = df[df['is_daylight']].copy()
        df_night = df[~df['is_daylight']].copy()


        if self.config.visuals.temperature_graph.daynight in ["day", 'both']:
            results['day'] = self.analyze_temperature_behavior_for_period(df=df_day,  period='day', results=results, polynomial_degree=2)
            # self.compare_polynomial_models(df=df_day, period='day')
        
        # results = self.analyze_temperature_behavior_glmm_by_period_MULTITHREAD(df=df, results=results)

        if self.config.visuals.temperature_graph.daynight in ["night", 'both']:
            results['night'] = self.analyze_temperature_behavior_for_period(df=df_night, period='night', results=results, polynomial_degree=2)

        # if self.config.visuals.temperature_graph.daynight == 'both':
        #     results['both'] = self.analyze_temperature_behavior_for_period(df=df,  period='both', results=results)
        
        # results = self.analyze_temperature_behavior_glmm(df=df, period='night', results=results)

        # results = self.analyze_temperature_behavior_glmm(df=df, results=results)

        # Create and print an interpretable table
        # table_unicode = self._create_interpretable_table(results, format_type="unicode")

        # print(results)
        table_unicode = self._create_comparison_table_compact(results, format_type="unicode")
        print("\nINTERPRETABLE MODEL RESULTS:")
        print(table_unicode)
        
        # Export table to Excel if needed
        if self.config.visuals.temperature_graph.export_excel:
            self._export_table_to_excel_compact(results)
            print("Results exported to Excel successfully.")

        return 0
